import sys
import os
from win32com.client import Dispatch
import tkinter
from tkinter import messagebox
from tkinter import filedialog
import openpyxl
from openpyxl.styles import Font
import datetime
import traceback
from subprocess import Popen
from cryptography.fernet import Fernet
from base64 import b64encode

from PyQt4 import QtCore, QtGui
from mm_gui import Ui_Widget
from mm_test import *
from email_contract import send_mail

# todo: Try to get new worksheets in spreadsheet to be generated with correct formatting (from template?)
# todo: split RevLine info into table entries instead of single long string. RevDate1, RevBy1, RevCheck1, RevCheckBy1?


class MyForm(QtGui.QWidget):

    def __init__(self, parent=None):
        QtGui.QWidget.__init__(self, parent)
        self.ui = Ui_Widget()
        self.ui.setupUi(self)
        self.setWindowTitle('LHS Inhouse Contract Tool')
        width = QtGui.QDesktopWidget().availableGeometry().width()
        height = QtGui.QDesktopWidget().availableGeometry().height()

        self.setGeometry(.25*width, .04*height, 0.5*width, 0.95*height)
        self.ui.tabWidget.tabBar().moveTab(1, 2)  # This re-orders the tabs. Maybe find a better way.

        self.ui.pushButton.clicked.connect(self.merge)

        self.ui.calendarWidget.activated.connect(self.show_date)
        self.ui.calendarWidget.clicked.connect(self.show_date)

        self.ui.calendarWidget_2.activated.connect(self.rev_show_date)
        self.ui.calendarWidget_2.clicked.connect(self.rev_show_date)

        program_prices = self.new_contract_fields()[1]
        for price in program_prices:
            price.editingFinished.connect(self.cost_calc)

        self.ui.pushButton_3.clicked.connect(self.browse_rev)

        self.ui.pushButton_4.clicked.connect(self.check_or_rev)

        self.ui.pushButton_5.clicked.connect(self.staff_section_file_open)

        self.ui.pushButton_6.clicked.connect(self.staff_section_file_write)

        self.ui.pushButton_7.clicked.connect(self.cancel_contract_file_open)

        self.ui.pushButton_8.clicked.connect(self.cancel_contract)

        rev_pgm_prices = self.rev_fields()[6]
        for price in rev_pgm_prices:
            price.editingFinished.connect(self.rev_cost_calc)

        rev_top_buttons = self.rev_fields()[9]
        for button in rev_top_buttons:
            button.clicked.connect(self.check_or_rev_text)

        global rev_program_indices
        rev_program_indices = [0]*10  # Initialize global variable to hold rev tab program indices

        global new_program_indices
        new_program_indices = None

        # Set path variable for either normal operation or testing
        global path
        path = '//Lhs-pfs01.berkeley.edu/lhs/B2H-EMS/MailMergeTest/'
        # path = './'

        # Section to set global variables for program titles and rooms available for each program. These will be read
        # in from an external Excel spreadsheet so that the values can easily be edited.
        root = tkinter.Tk()
        root.withdraw()
        try:
            global program_titles, program_rooms, program_times
            programs_wb = openpyxl.load_workbook('LHS_programs_list.xlsx')
            programs_ws = programs_wb.active
            counter, none_flag, program_titles, program_rooms = 0, 0, [], []
            for row in programs_ws.iter_rows('A2:B100'):
                for cell in row:
                    if cell.value is None:
                        none_flag += 1
                    elif counter % 2 == 0:  # even numbers on counter are A column in spreadsheet, program titles
                        program_titles.append(cell.value)
                        none_flag = 0  # reset blank cell counter to zero at the beginning of each row
                    else:  # odd numbers on counter are B column in spreadsheet, program rooms
                        program_rooms.append([item.strip() for item in str.split(str(cell.value), ',')])
                    counter += 1
                if none_flag == 2:  # if we hit two blank cells in a single row, stop reading
                    break
            if len(program_titles) != len(program_rooms):
                tkinter.messagebox.showerror("Error!", "Error reading LHS_programs_list.xlsx\n"
                                                       "Make sure every program has available rooms listed.")
                raise Exception
            program_titles.append('Custom Program')
            program_times = [''] + [item.strip() for item in str.split(programs_ws['C2'].value, ',')]
        except Exception as err:
            tkinter.messagebox.showerror("Error!", err)
            raise Exception

        new_program_boxes = self.new_contract_fields()[2]
        for box in new_program_boxes:
            box.clear()  # delete selections entered in QtCreator
            box.addItems([''] + program_titles)  # add selections read in from LHS_programs_list.xlsx
            box.activated.connect(self.room_list)

        rev_pgm_boxes = self.rev_fields()[5]
        for box in rev_pgm_boxes:
            box.clear()  # delete selections from QtCreator
            box.addItems([''] + program_titles)  # adds workshops read in from external spreadsheet
            box.activated.connect(self.rev_box_update)
            # box.activated.connect(self.rev_program_box_select)

        new_time_boxes = self.new_contract_fields()[0]
        for box in new_time_boxes:
            box.clear()  # delete selections from QtCreator
            box.addItems(program_times + ['Custom'])  # add selections read in from program list spreadsheet
            box.activated.connect(self.time_box_update)

        rev_time_boxes = self.rev_fields()[2]
        for box in rev_time_boxes:
            box.clear()  # delete selections from QtCreator
            box.addItems(program_times + ['Custom'])  # add selections read in from program list spreadsheet
            box.activated.connect(self.rev_time_box_update)

    @staticmethod
    def contract_info():
        global program_titles, program_rooms, program_times

        pgm_titles = program_titles
        room_groups = [['140', '141'], ['120 (20 max)', '140', '141', 'Niche (16 max)'], ['155'],
                       ['140', '145', '155 (not Fri if Bubbles)', '165'],
                       ['120 (if small group)', '140', '145', '155', '165'],
                       ['140', '145', '155 (not Fri if Bubbles)', '165', '175 (last resort)'],
                       ['145', '155 (not Fri if Bubbles)', '165 (last resort)'], ['LHS Museum'], ['LHS Auditorium'],
                       ['120 (20 max)', '140', '141', '145', '155', 'Niche (16 max)'], ['145', '165'], ['Planetarium'],
                       ['145', '155 (not Fri if Bubbles)'], ['140', '141', '145', '155', '165'],
                       ['140', '145', '155', '165'], ['145', '155 (not Fri if Bubbles)', '165'], ['175']]
        room_group_map = [0, 0, 1, 2, 3, 16, 0, 4, 16, 16, 16, 5, 1, 6, 3, 3, 7, 7, 3, 1, 15, 9, 10, 15, 3, 16, 16, 11,
                          12, 11, 3, 13, 16, 16, 14, 16, 16, 3, 8, 8]
        merge_keys = ["FirstPgmTime", "SecondPgmTime", "ThirdPgmTime", "FourthPgmTime", "FifthPgmTime", "SixthPgmTime",
                      "SeventhPgmTime", "EighthPgmTime", "NinthPgmTime", "TenthPgmTime",
                      "FirstPgmTitle", "SecondPgmTitle", "ThirdPgmTitle", "FourthPgmTitle", "FifthPgmTitle",
                      "SixthPgmTitle", "SeventhPgmTitle", "EighthPgmTitle", "NinthPgmTitle", "TenthPgmTitle",
                      "ContractEMS", "BookedBy", "DateTaken", "SchoolName", "SchoolFax", "SchoolAddress", "SchoolPhone",
                      "ReserverName", "ReserverPhone", "ReserverEmail", "ContactName", "ContactPhone", "DayDate",
                      "RevBy",
                      "FirstPgmGrades", "SecondPgmGrades", "ThirdPgmGrades", "FourthPgmGrades", "FifthPgmGrades",
                      "SixthPgmGrades", "SeventhPgmGrades", "EighthPgmGrades", "NinthPgmGrades", "TenthPgmGrades",
                      "FirstPgmStudents", "SecondPgmStudents", "ThirdPgmStudents", "FourthPgmStudents",
                      "FifthPgmStudents", "SixthPgmStudents", "SeventhPgmStudents", "EighthPgmStudents",
                      "NinthPgmStudents", "TenthPgmStudents", "FirstPgmPrice", "SecondPgmPrice", "ThirdPgmPrice",
                      "FourthPgmPrice", "FifthPgmPrice", "SixthPgmPrice", "SeventhPgmPrice", "EighthPgmPrice",
                      "NinthPgmPrice", "TenthPgmPrice", "TotalCost", "PayDate",
                      "SpecialConsiderations", "RevNote"]
        # non_custom_times = ['', '10:00-11:00', '11:10-12:10', '12:30-1:30', '1:40-2:40']
        standard_times = program_times
        payment_info_keys = ['PONumber', 'POAmount', 'PORecDate', 'CheckCC1', 'CheckCCRecDate1', 'CheckCCAmount1',
                             'CheckCCRec1', 'CheckCC2', 'CheckCCRecDate2', 'CheckCCAmount2', 'CheckCCRec2', 'CheckCC3',
                             'CheckCCRecDate3', 'CheckCCAmount3', 'CheckCCRec3']
        pgm_rooms = program_rooms
        # xl_entries = ["ContractEMS", "SchoolName", "SchoolAddress", "SchoolCity", "SchoolZip", "SchoolPhone",
        #               "ReserverName", "ContactName", "ContactPhone", "ContactFax", "SpecialConsiderations",
        #               "ReserverEmail", "PgmDate", "PgmDay", "ProgramTitle", "PgmTime", "Grades", "Students", "Room",
        #               "CostForPgm", "TOTALCOST", "DueDate", "BookedBy"]

        return [pgm_titles, room_groups, room_group_map, merge_keys, standard_times, payment_info_keys, pgm_rooms]

    def new_contract_fields(self):  # Declare all UI fields on New Contract tab
        time_boxes = [self.ui.comboBox_21, self.ui.comboBox_22, self.ui.comboBox_23, self.ui.comboBox_24,
                      self.ui.comboBox_25, self.ui.comboBox_26, self.ui.comboBox_27, self.ui.comboBox_28,
                      self.ui.comboBox_29, self.ui.comboBox_30]
        price_fields = [self.ui.lineEdit_13, self.ui.lineEdit_17, self.ui.lineEdit_21, self.ui.lineEdit_25,
                        self.ui.lineEdit_29, self.ui.lineEdit_33, self.ui.lineEdit_37, self.ui.lineEdit_41,
                        self.ui.lineEdit_45, self.ui.lineEdit_49]
        title_boxes = [self.ui.comboBox, self.ui.comboBox_3, self.ui.comboBox_5, self.ui.comboBox_7,
                       self.ui.comboBox_9, self.ui.comboBox_11, self.ui.comboBox_13, self.ui.comboBox_15,
                       self.ui.comboBox_17, self.ui.comboBox_19]
        room_boxes = [self.ui.comboBox_2, self.ui.comboBox_4, self.ui.comboBox_6, self.ui.comboBox_8,
                      self.ui.comboBox_10, self.ui.comboBox_12, self.ui.comboBox_14, self.ui.comboBox_16,
                      self.ui.comboBox_18, self.ui.comboBox_20]
        grade_fields = [self.ui.lineEdit_11, self.ui.lineEdit_15, self.ui.lineEdit_19, self.ui.lineEdit_23,
                        self.ui.lineEdit_27, self.ui.lineEdit_31, self.ui.lineEdit_35, self.ui.lineEdit_39,
                        self.ui.lineEdit_43, self.ui.lineEdit_47]
        student_fields = [self.ui.lineEdit_12, self.ui.lineEdit_16, self.ui.lineEdit_20, self.ui.lineEdit_24,
                          self.ui.lineEdit_28, self.ui.lineEdit_32, self.ui.lineEdit_36, self.ui.lineEdit_40,
                          self.ui.lineEdit_44, self.ui.lineEdit_48]
        top_fields = [self.ui.lineEdit_3, self.ui.lineEdit_52, self.ui.lineEdit_10, self.ui.lineEdit_4,
                      self.ui.lineEdit_2, self.ui.lineEdit_5, self.ui.lineEdit_108, self.ui.lineEdit_109,
                      self.ui.lineEdit_8, self.ui.lineEdit_6,
                      self.ui.lineEdit_7, self.ui.lineEdit_9, self.ui.lineEdit, self.ui.lineEdit_54,
                      self.ui.lineEdit_53, self.ui.lineEdit_51]

        return [time_boxes, price_fields, title_boxes, room_boxes, grade_fields, student_fields, top_fields]

    def rev_fields(self):
        # top fields order: ContractEMS, BookedBy, DateTaken, SchoolName, SchoolFax, SchoolAddress, SchoolPhone,
        #                   ReserverName, ReserverPhone, ReserverEmail, ContactName, ContactPhone,
        #                   Program date (day and date), RevBy
        rev_top_fields = [self.ui.lineEdit_76, self.ui.lineEdit_77, self.ui.lineEdit_87, self.ui.lineEdit_80,
                          self.ui.lineEdit_81, self.ui.lineEdit_82, self.ui.lineEdit_110, self.ui.lineEdit_111,
                          self.ui.lineEdit_83, self.ui.lineEdit_78,
                          self.ui.lineEdit_79, self.ui.lineEdit_86, self.ui.lineEdit_84, self.ui.lineEdit_85,
                          self.ui.lineEdit_90, self.ui.lineEdit_99]
        rev_note_fields = [self.ui.textEdit_2, self.ui.textEdit_3]
        rev_time_boxes = [self.ui.comboBox_37, self.ui.comboBox_39, self.ui.comboBox_49, self.ui.comboBox_48,
                          self.ui.comboBox_42, self.ui.comboBox_52, self.ui.comboBox_56, self.ui.comboBox_55,
                          self.ui.comboBox_34, self.ui.comboBox_32]
        rev_grade_fields = [self.ui.lineEdit_34, self.ui.lineEdit_22, self.ui.lineEdit_64, self.ui.lineEdit_61,
                            self.ui.lineEdit_30, self.ui.lineEdit_65, self.ui.lineEdit_69, self.ui.lineEdit_71,
                            self.ui.lineEdit_59, self.ui.lineEdit_56]
        rev_student_fields = [self.ui.lineEdit_14, self.ui.lineEdit_18, self.ui.lineEdit_60, self.ui.lineEdit_62,
                              self.ui.lineEdit_67, self.ui.lineEdit_68, self.ui.lineEdit_66, self.ui.lineEdit_72,
                              self.ui.lineEdit_46, self.ui.lineEdit_57]
        rev_pgm_boxes = [self.ui.comboBox_41, self.ui.comboBox_40, self.ui.comboBox_46, self.ui.comboBox_43,
                         self.ui.comboBox_50, self.ui.comboBox_60, self.ui.comboBox_54, self.ui.comboBox_58,
                         self.ui.comboBox_35, self.ui.comboBox_33]
        rev_pgm_prices = [self.ui.lineEdit_26, self.ui.lineEdit_42, self.ui.lineEdit_63, self.ui.lineEdit_38,
                          self.ui.lineEdit_74, self.ui.lineEdit_75, self.ui.lineEdit_70, self.ui.lineEdit_73,
                          self.ui.lineEdit_58, self.ui.lineEdit_55]
        rev_room_boxes = [self.ui.comboBox_38, self.ui.comboBox_45, self.ui.comboBox_47, self.ui.comboBox_44,
                          self.ui.comboBox_51, self.ui.comboBox_57, self.ui.comboBox_53, self.ui.comboBox_59,
                          self.ui.comboBox_36, self.ui.comboBox_31]
        rev_bottom_fields = [self.ui.lineEdit_88, self.ui.lineEdit_89]
        check_or_rev_buttons = [self.ui.radioButton_3, self.ui.radioButton_4]

        return [rev_top_fields, rev_note_fields, rev_time_boxes, rev_grade_fields, rev_student_fields, rev_pgm_boxes,
                rev_pgm_prices, rev_room_boxes, rev_bottom_fields, check_or_rev_buttons]

    def check_or_rev_text(self):  # This changes label text for either contract checking or revising
        if self.ui.radioButton_3.isChecked():
            self.ui.label_75.setText('Checked By:')
            self.ui.lineEdit_76.setReadOnly(False)
        else:
            self.ui.label_75.setText('Revised By:')
            self.ui.lineEdit_76.setReadOnly(True)

    def check_or_rev(self):  # Check radio button values to decide if checking or revising, then run correct module
        if self.ui.radioButton_3.isChecked():
            self.check_contract()
        else:
            self.apply_rev()

    @staticmethod
    def creation_date(filename):
        # This has been changed to look for the modified date instead of creation date since contract files might be
        # moved to different folder which resets the creation date
        t = os.path.getmtime(filename)
        return datetime.datetime.fromtimestamp(t).strftime('%b%Y')

    @staticmethod
    def contract_kp_builder(template_file, data_file):

        # Read contract template and grab all field names
        file_xml = read_docx(template_file)
        my_etree = etree.fromstring(file_xml)
        keys = []
        for node in my_etree.iter(tag=etree.Element):
            if check_element_is(node, 'fldChar'):  # Once we've hit this, we're money...
                # Now, we're looking for this attribute: w:fldCharType="separate"
                if node.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}fldCharType') == "separate":
                    node_value = node.getparent().getnext().getchildren()[1].text
                    node_value = node_value[1:-1]
                    keys.append(node_value)
            elif check_element_is(node, 'fldSimple'):  # Once we've hit this, we're money...
                node_value = node.getchildren()[0].getchildren()[1].text
                node_value = node_value[1:-1]
                keys.append(node_value)

        # Read contract data file and grab all field values
        file_xml2 = read_docx(data_file)
        my_etree2 = etree.fromstring(file_xml2)
        values = []
        for node in my_etree2.iter(tag=etree.Element):
            if check_element_is(node, 'fldChar'):  # Once we've hit this, we're money...
                # Now, we're looking for this attribute: w:fldCharType="separate"
                if node.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}fldCharType') == "separate":
                    node_value = node.getparent().getnext().getchildren()[1].text
                    values.append(node_value)
            elif check_element_is(node, 'fldSimple'):  # Once we've hit this, we're money...
                node_value = node.getchildren()[0].getchildren()[1].text
                values.append(node_value)

        new_kp = {k: v for k, v in zip(keys, values)}

        return new_kp

    @staticmethod
    def save_as_pdf(infile, outfile):
        wdFormatPDF = 17

        in_file = os.path.abspath(infile)
        out_file = os.path.abspath(outfile)

        word = Dispatch('Word.Application')
        doc = word.Documents.Open(in_file)
        doc.SaveAs(out_file, FileFormat=wdFormatPDF)
        doc.Close()
        word.Quit()

        return

    def cancel_contract_file_open(self):
        global path
        root = tkinter.Tk()
        root.withdraw()

        file_path = tkinter.filedialog.askopenfilename(initialdir=path)  # this gives full path

        if file_path == '':
            return
        elif self.valid_contract_file_check(file_path) == 1:
            tkinter.messagebox.showerror("Error!", "Invalid contract file")
            return
        elif file_path[-7:-5] == 'CX':
            tkinter.messagebox.showerror("Error!", "This contract has already been canceled")
            return
        elif os.path.exists(file_path[:-5]+'_CX.docx') or os.path.exists(file_path[:-5]+'_REV_CX.docx'):
            tkinter.messagebox.showerror("Error!", "A canceled version of this contract already exists")
            return
        elif os.path.exists(file_path[:-5]+'_REV.docx'):
            tkinter.messagebox.showerror("Error!", "A revised version of this contract exists.\n"
                                                   "Cancellation must be done on newest version.")
            return

        self.ui.lineEdit_100.setText(os.path.abspath(file_path))
        self.ui.lineEdit_107.clear()  # clear initials each time a new file is selected
        return

    def search_whole_spreadsheet(self, date_from_file, record_number):
        """
        In the event that a contract's modification date is not pointing to the right spreadsheet month/year, this
        function will simply cycle backwards through worksheets and possibly spreadsheet file years until the correct
        contract number is found.
        :param date_from_file: string date in '%b%Y' format where file date claims record should be
        :param record_number: contract number for record you're seeking
        :return: record_arr: an array of spreadsheet cell values for the record
        :return: found_sheet_name: string date in format '%b%Y' of worksheet name where record was found
        """
        global path
        months = ["Dec", "Nov", "Oct", "Sep", "Aug", "Jul", "Jun", "May", "Apr", "Mar", "Feb", "Jan"]
        year_from_file = date_from_file[-4:]
        spreadsheet_path = os.path.normpath(os.path.join(path,
                                                         "inhouse_contract_spreadsheet_"+year_from_file+".xlsx"))
        month_ind = [i for i, m in enumerate(months) if months[i] in date_from_file][0]
        record_arr = None
        while month_ind < 12:
            try:
                print("Checking", spreadsheet_path, "on sheet", months[month_ind]+year_from_file)
                record_arr = self.get_xl_record(spreadsheet_path, months[month_ind]+year_from_file, record_number,
                                                search_whole_call=True)
            except FileNotFoundError as err:
                tkinter.messagebox.showerror("Error!", "Attempted to read a spreadsheet that doesn't exist")
                break
            if not record_arr:
                if month_ind == 11:
                    year_from_file = str(int(year_from_file) - 1)
                    spreadsheet_path = spreadsheet_path[:-9] + year_from_file + ".xlsx"
                    month_ind = 0
                else:
                    month_ind += 1
            else:
                break
        found_sheet_name = months[month_ind]+year_from_file
        return record_arr, found_sheet_name

    def cancel_contract(self):
        global path
        # This needs to save the contract file with _CX added to filename, delete the original file, and add
        # strike-through style to every cell of its excel record
        confirm = tkinter.messagebox.askokcancel("Confirm cancellation", "Click OK to confirm cancellation")
        if confirm != 1:
            tkinter.messagebox.showinfo("Aborted", "Contract cancellation aborted")

        # Get the absolute path for the contract file from the lineEdit field on the GUI
        file_path = self.ui.lineEdit_100.text()
        current_folder = file_path.strip(file_path.split('\\')[-1])

        # Get all the merge field values from the contract file. Note the contract has already been checked for
        # validity when the file was selected.
        contract_kp = self.contract_kp_builder('merge_template.docx', file_path)

        # Get initials of person canceling contract
        canceled_by = self.ui.lineEdit_107.text()
        if canceled_by == '':
            tkinter.messagebox.showerror("Error!", "Please enter your initials")
            return
        canceled_date = datetime.datetime.now().strftime("%#m/%#d/%Y")
        contract_kp.update({"Canceled": 'CANCELED '+canceled_date+' '+canceled_by})

        # Section to edit excel record. Find record in workbook, iterate through each cell setting
        # Font(strikethrough=True)
        # need original creation date of contract
        if file_path[-8:-5] == 'REV':
            try:
                orig_file = [file for file in os.listdir(current_folder) if contract_kp["ContractEMS"] in file and
                             file[-8:-5] not in ['REV', '_CX']]
                if not orig_file:  # if list returns empty, give error and return
                    tkinter.messagebox.showerror("Error!", "No un-revised version of this contract was found")
                    return
                orig_path = os.path.join(current_folder, orig_file[0])
                # if "DateWritten" not in contract_kp or not contract_kp["DateWritten"]:
                #     dw = datetime.datetime.fromtimestamp(os.path.getmtime(orig_path)).strftime('%m/%d/%Y')
                #     contract_kp.update({"DateWritten": dw})
                c_date = self.creation_date(orig_path)
                _, found_sheet = self.search_whole_spreadsheet(c_date, contract_kp["ContractEMS"])
                if found_sheet != c_date:
                    c_date = found_sheet
                # Check if DateWritten timestamp exists in contract. If so, and if it doesn't point to the correct
                # spreadsheet page, remove it from the merging dict
                # FIXME
                try:
                    if "DateWritten" in contract_kp and contract_kp["DateWritten"]:
                        date_for_sheet = datetime.datetime.strptime(contract_kp["DateWritten"], '%m/%d/%Y').strftime('%b%Y')
                        if date_for_sheet != found_sheet:
                            contract_kp.update({"DateWritten": None})
                except:
                    pass

            except Exception as err:
                tkinter.messagebox.showerror("Error!", str(err) + ' \nCrash log written to error_log.txt')
                old_stdout = sys.stdout
                log_file = open('error_log.txt', 'w')
                sys.stdout = log_file
                traceback.print_exc()
                sys.stdout = old_stdout
                log_file.close()
                return
        else:

            # if "DateWritten" not in contract_kp or not contract_kp["DateWritten"]:
            #     dw = datetime.datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%m/%d/%Y')
            #     contract_kp.update({"DateWritten": dw})
            c_date = self.creation_date(file_path)
            _, found_sheet = self.search_whole_spreadsheet(c_date, contract_kp["ContractEMS"])
            if found_sheet != c_date:
                c_date = found_sheet
            # Check if DateWritten timestamp exists in contract. If so, and if it doesn't point to the correct
            # spreadsheet page, remove it from the merging dict
            # FIXME
            try:
                if "DateWritten" in contract_kp and contract_kp["DateWritten"]:
                    date_for_sheet = datetime.datetime.strptime(contract_kp["DateWritten"], '%m/%d/%Y').strftime('%b%Y')
                    if date_for_sheet != found_sheet:
                        contract_kp.update({"DateWritten": None})
            except:
                pass

        # Append _CX to filename
        out_path = file_path[:-5]+'_CX.docx'

        try:
            wb = openpyxl.load_workbook(path+'inhouse_contract_spreadsheet_'+c_date[-4:]+'.xlsx')
            ws = wb.get_sheet_by_name(c_date)
        except FileNotFoundError:
            tkinter.messagebox.showerror("Error!", "Spreadsheet not found")
            return
        except KeyError:
            tkinter.messagebox.showerror("Error!", "Worksheet not found in spreadsheet")
            return

        record_start = []
        break_flag = 0
        for row in ws.iter_rows('A2:A500'):
            for cell in row:
                if cell.value == contract_kp["ContractEMS"]:
                    record_start = str.split(str(cell), '.')[1][:-1]  # this gives cell letter and number as string
                    break_flag = 1
            if break_flag == 1:  # stop after finding first instance (duplicate entries not allowed in spreadsheet)
                break
        # Make sure record was found
        if record_start is None:
            tkinter.messagebox.showerror("Error!", "Contract record not found in spreadsheet")
            return

        # Find record column length by scanning program title row. Maximum of 10 programs are possible.
        record_end = []
        break_flag = 0
        for row in ws.iter_rows('P'+record_start[1:]+':P'+str(int(record_start[1:]) + 10)):
            for cell in row:
                if cell.value is None:
                    # this madness gives the string version of the row just BEFORE the one where blank cell was found
                    record_end = 'AD' + str(int(str.split(str(cell), '.')[1][1:-1]) - 1)
                    break_flag = 1
            if break_flag == 1:
                break

        # Confirm decision to cancel
        ok = tkinter.messagebox.askokcancel("Confirm file name", "File being saved as: "+os.path.split(out_path)[1] +
                                            "\nAre you really, really sure?")
        if ok != 1:
            tkinter.messagebox.showinfo("Aborted", "Contract cancellation aborted")
            return

        # Set each row in the record to have the strikethrough style
        for row in ws.iter_rows(record_start+':'+record_end):
            for cell in row:
                if str.split(str(cell), '.')[1][:-1] == 'Y'+record_start[1:]:
                    cell.value = canceled_date+' '+canceled_by  # write date and initials of cancellation in workbook
                # elif str.split(str(cell), '.')[1][:-1] == 'Z'+record_start[1:]:
                #     pass
                else:
                    cell.font = Font(strikethrough=True)

        # Write spreadsheet record
        try:
            wb.save(filename=path+'inhouse_contract_spreadsheet_'+c_date[-4:]+'.xlsx')
        except PermissionError as err:
            tkinter.messagebox.showerror("Error!", "Permission error. Make sure spreadsheet is not open anywhere, "
                                                   "then try again.\n\n" + str(err))
            return

        # Merge word doc
        docxmerge('merge_template.docx', contract_kp, out_path)

        self.ui.lineEdit_100.clear()
        self.ui.lineEdit_107.clear()
        os.remove(file_path)
        tkinter.messagebox.showinfo("Success", "Contract successfully canceled")

        return

    @staticmethod
    def get_xl_record(workbook, sheet_name, record_number, search_whole_call=None):
        print('Grabbing Excel record...')
        print('Loading file: ' + workbook)
        print('Loading sheet: ' + sheet_name)
        root = tkinter.Tk()
        root.withdraw()
        xfile = openpyxl.load_workbook(workbook)
        sheet = xfile.get_sheet_by_name(sheet_name)
        record_start = []

        # Check that excel file exists and worksheet exists in file
        if not os.path.exists(workbook):
            print('Workbook does not exist')
            tkinter.messagebox.showerror("Error!", "Workbook does not exist")
            return
        elif sheet_name not in xfile.sheetnames:
            print('Could not find worksheet in file')
            tkinter.messagebox.showerror("Error!", "Could not find worksheet in workbook")
            return

        # Scan through first column of spreadsheet to find record number
        break_flag = 0
        for row in sheet.iter_rows('A2:A500'):
            for cell in row:
                # print(cell.value)
                if cell.value == record_number:
                    record_start = str.split(str(cell), '.')[1][:-1]
                    break_flag = 1
                    # print('Record begins at:', record_start)
                if break_flag == 1:
                    break

        # Check if record number was not found
        if not record_start:
            print('Record number not found in spreadsheet')
            if not search_whole_call:
                tkinter.messagebox.showerror("Error!", "Record number not found in workbook")
            return

        # Starting at row of LAST instance of record number, scan through program title column to find record length
        for row in sheet.iter_rows('P' + record_start[1:] + ':P' + str(int(record_start[1:]) + 10)):
            for cell in row:
                # print(cell.value)
                if cell.value is None:
                    # todo: does this need to be changed back to column X to avoid breaking anything?
                    record_end = 'AD' + str(int(str.split(str(cell), '.')[1][1:-1]) - 1)
                    # print('Record ends at:', record_end)
                    record_array = [[cell.value for cell in tuple(sheet[record_start:record_end])[i]] for
                                    i in range(int(record_end[2:]) - int(record_start[1:]) + 1)]
                    return record_array

    def build_xl_record(self, kp, room_numbers, *rev_or_cx):
        pgm_titles = []
        pgm_times = []
        pgm_grades = []
        pgm_students = []
        pgm_prices = []
        room_numbers = [str.split(room_numbers[i], '(')[0].strip() for i in range(len(room_numbers))
                        if room_numbers[i] is not None]

        for i in range(len(room_numbers)):
            pgm_titles.append(kp[self.contract_info()[3][i+10]])
            pgm_times.append(kp[self.contract_info()[3][i]])
            pgm_grades.append(kp[self.contract_info()[3][i+34]])
            pgm_students.append(kp[self.contract_info()[3][i+44]])
            pgm_prices.append(kp[self.contract_info()[3][i+54]])

        split_address = [item.strip() for item in str.split(kp["SchoolAddress"], ',')]
        xl_record = [kp["ContractEMS"], kp["SchoolName"], split_address[0], split_address[1], split_address[2],
                     kp["SchoolPhone"], kp["ReserverName"], kp["ReserverPhone"], kp["ContactName"], kp["ContactPhone"],
                     kp["SchoolFax"], kp["SpecialConsiderations"], kp["ReserverEmail"],
                     str.split(kp["DayDate"], ',')[1].strip(), str.split(kp["DayDate"], ',')[0],
                     pgm_titles, pgm_times, pgm_grades, pgm_students,
                     room_numbers, pgm_prices, kp["TotalCost"], kp["PayDate"], kp["BookedBy"]]

        if rev_or_cx[0] == 'rev':  # if this record is a revision (or cancellation), record date and note
            print('Writing additional revision cells')
            xl_record.append(datetime.datetime.now().strftime("%#m/%#d/%Y")+' '+kp["RevBy"])
            xl_record.append(kp["RevNote"])
        elif rev_or_cx[0] == 'not_rev':
            xl_record.append('')
            xl_record.append('')
        # elif rev_or_cx[0] == 'rev_check':
        #     xl_record.append('')
        #     xl_record.append(kp["RevNote"])

        # Add payment info columns
        try:
            po = kp["PONumber"]+', '+kp["POAmount"]+', '+kp["PORecDate"]
        except TypeError:
            po = None
        try:
            check_cc_1 = kp["CheckCC1"]+', '+kp["CheckCCRecDate1"]+', '+kp["CheckCCAmount1"]+', '+kp["CheckCCRec1"]
        except TypeError:
            check_cc_1 = None
        try:
            check_cc_2 = kp["CheckCC2"]+', '+kp["CheckCCRecDate2"]+', '+kp["CheckCCAmount2"]+', '+kp["CheckCCRec2"]
        except TypeError:
            check_cc_2 = None
        try:
            check_cc_3 = kp["CheckCC3"]+', '+kp["CheckCCRecDate3"]+', '+kp["CheckCCAmount3"]+', '+kp["CheckCCRec3"]
        except TypeError:
            check_cc_3 = None
        for item in [po, check_cc_1, check_cc_2, check_cc_3]:
            xl_record.append(item)
        # xl_record.append([item for item in [po, check_cc_1, check_cc_2, check_cc_3]])

        # Pad array with None to make rectangular
        for i in range(len(xl_record)):
            if i in range(15) or i in range(21, len(xl_record)):
                xl_record[i] = [xl_record[i]] + [None]*(len(room_numbers) - 1)
        # print(pgm_titles)
        # print(xl_record)
        return xl_record

    @staticmethod
    def write_xl_record(workbook, sheet_name, record_number, new_xl_record, **kwargs):
        """This method writes excel record when changing/revising contracts"""

        xfile = openpyxl.load_workbook(workbook)
        sheet = xfile.get_sheet_by_name(sheet_name)
        record_start = []
        record_end = []
        break_flag = 0

        # Find record in spreadsheet
        for row in sheet.iter_rows('A2:A500'):  # Check for ALL instances of record number in first column
            for cell in row:
                if cell.value == record_number:
                    record_start = str.split(str(cell), '.')[1][:-1]  # store cell where record number is found
                    break_flag = 1
                    break
            if break_flag == 1:
                break

        # Check to see if contract data was not found in spreadsheet
        if record_start is None:
            print('Could not find record number in spreadsheet')
            return

        # Iterate through program title column, starting at row where record number was found, to see how many rows
        # the record occupies
        break_flag = 0
        for row in sheet.iter_rows('P' + record_start[1:] + ':P' + str(int(record_start[1:]) + 12)):
            for cell in row:
                if cell.value is None:
                    record_end = 'AD' + str(int(str.split(str(cell), '.')[1][1:-1]) - 1)
                    break_flag = 1
                    break
            if break_flag == 1:
                break
        # print('Existing record goes from ' + record_start + ' to ' + record_end)
        if not record_end:
            tkinter.messagebox.showerror("Error!", "Could not find end of record in spreadsheet")
            return

        new_booked_rooms = [new_xl_record[19][i] for i in range(len(new_xl_record[19]))]  # Gives array column length

        # Check for record length difference, i.e. if programs have been added or removed
        length_diff = len(new_xl_record[19]) - int(record_end[2:]) + int(record_start[1:]) - 1
        # if length_diff == 0:
        #     print('New record is SAME LENGTH as existing record.')
        first_blank_row = []
        if length_diff != 0:
            print('Length difference is:', length_diff)
            # Scan program title row to find first instance of two consecutive blank cells (end of sheet data)
            blank_counter = 0
            break_flag = 0
            for row in sheet.iter_rows('P' + str(int(record_end[2:]) + 1) + ':P500'):
                for cell in row:
                    if cell.value is None:
                        blank_counter += 1
                    else:
                        blank_counter = 0
                    if blank_counter > 1:
                        first_blank_row = int(str.split(str(cell), '.')[1][1:-1]) - 1
                        break_flag = 1
                        break
                if break_flag == 1:
                    break
            if not first_blank_row:
                tkinter.messagebox.showerror("Error!", "Could not find first blank row after record")
                return
            # print('Found first blank row after all sheet data at row:', first_blank_row)
            if first_blank_row == int(record_end[2:]) + 1 and length_diff < 0:
                # print('This record is the last entry in the spreadsheet, but programs have been removed.')
                # print('Deleting removed programs...')
                row = int(record_end[2:]) + 1 + length_diff
                cols = 'PQRSTU'
                for i in range(abs(length_diff)):
                    for j in range(6):
                        sheet[cols[j]+str(row)].value = None
                    row += 1
                # print('Done!')
            else:
                # Store all data in spreadsheet from end of current record to end of sheet so it can be shifted

                # This is a bit ugly, but it produces a list of ROWS, where each ROW is a list of CELLS that make up
                # that row
                data_store = [[[cell[i].value for cell in sheet.iter_rows('A'+str(row)+':AD'+str(row))][0] for i in
                               range(30)] for row in range(int(record_end[2:]) + 1, first_blank_row)]
                # print('Data stored:', data_store)

                if length_diff > 0:
                    print('Shifting data down', length_diff, 'row(s)...')
                else:
                    # Check to make sure we only shift UP fewer rows than the original record occupied
                    if abs(length_diff) >= int(record_end[2:]) - int(record_start[1:]) + 1:
                        root = tkinter.Tk()
                        root.withdraw()
                        tkinter.messagebox.showerror("Error!", "Spreadsheet error. Process terminating.")
                        return
                    print('Shifting data up', abs(length_diff), 'row(s)...')
                row = int(record_end[2:]) + 1 + length_diff  # initialize row counter to correct location
                # cols = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                        'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']
                for i in range(len(data_store)):
                    for j in range(30):
                        sheet[cols[j]+str(row)].value = data_store[i][j]
                    row += 1
                print('Done!')
        else:
            print('New record is SAME LENGTH as existing record.')

        # Write new record into spreadsheet
        if 'yes' in kwargs.values():  # yes value means contract was previously revised, so don't overwrite rev cells
            last_col = 'X'
            last_col_ind = 23
        else:  # case where contract has not been revised, so overwrite all cells
            last_col = 'Z'
            last_col_ind = 25

        print('Writing new Excel data...')
        record_index = [0, 0]
        for row in sheet.iter_rows(record_start + ':AD' + str(int(record_start[1:]) + len(new_booked_rooms) - 1)):
            for cell in row:
                # print('Column counter:', record_index[0], 'Row counter:', record_index[1])
                # Check for revision flag to see if column Y should be overwritten. If 'yes' is in kwargs.values(),
                # that indicates that a previous revision has been done and should NOT have its initials and date
                # overwritten, i.e. THIS writing process is not a NEW revision, but needs to retain information about
                # the previous revision. In any case, column Z (which holds the revision note) should be written. If
                # there is no revision note, writing a blank entry should be fine.
                if str.split(str(cell), '.')[1][0] == 'Y' and 'yes' in kwargs.values():
                    pass
                else:
                    if type(new_xl_record[record_index[0]]) is list:  # check if current column has more than one entry
                        cell.value = new_xl_record[record_index[0]][record_index[1]]
                        # print('Writing', new_xl_record[record_index[0]][record_index[1]])
                    else:
                        cell.value = new_xl_record[record_index[0]]
                        # print('Writing', new_xl_record[record_index[0]])
                if record_index[0] == 29:
                    record_index[0] = 0  # reset column counter once the end of row is reached
                else:
                    record_index[0] += 1  # increment column counter
            record_index[1] += 1  # increment row counter

        xfile.save(workbook)
        print('Done!')

        return

    def rev_cost_calc(self):
        price_fields = self.rev_fields()[6]
        total = 0
        for field in price_fields:
            if field.text() != '':
                total += int(field.text())
        self.ui.lineEdit_88.setText('$' + str('{:.2f}'.format(total)))
        return '{:.2f}'.format(total)

    def rev_box_update(self):
        # Anytime a program box value is selected, this function will check which program titles have changed and update
        # each corresponding room box to the correct choices
        global rev_program_indices

        program_indices = [box.currentIndex() for box in self.rev_fields()[5]]

        box_names_kp = {k: v for k, v in zip(self.rev_fields()[5], self.rev_fields()[7])}
        previous_pgm_indices = {k: v for k, v in zip(self.rev_fields()[5], rev_program_indices)}

        programs_and_rooms = {k: v for k, v in zip(self.contract_info()[0][:-1], self.contract_info()[6])}

        for key, value in box_names_kp.items():
            # print(key.currentIndex(), previous_pgm_indices[key])
            if key.currentIndex() != previous_pgm_indices[key]:
                if key.currentText() == '':
                    key.setEditable(False)
                    value.setEditable(False)
                    value.clear()
                # elif key.currentText() not in program_room_map_kp:
                elif key.currentText() not in programs_and_rooms:
                    key.setEditable(True)
                    value.clear()
                    value.setEditable(True)
                elif key.currentText() in programs_and_rooms:
                    key.setEditable(False)
                    value.setEditable(False)
                    value.clear()
                    # value.addItems(room_groups[program_room_map_kp[key.currentText()]])
                    value.addItems(programs_and_rooms[key.currentText()])

        rev_program_indices = program_indices  # update global variable rev_program_indices with current box indices
        return

    def rev_field_scraper(self):
        """Get values for all fields on contract revision tab so they can be merged into word doc"""

        # Get values for all comboBox fields and append them to values list
        time_boxes = self.rev_fields()[2]
        program_boxes = self.rev_fields()[5]
        boxes = [time_boxes, program_boxes]
        merge_values = []
        for box in boxes:
            for i in box:
                # print(i.currentText())
                if i.currentText() != '':
                    merge_values.append(i.currentText())
                else:
                    merge_values.append(None)

        # Get values for all lineEdit fields and append them to values list
        line_edit_fields = [self.rev_fields()[0], self.rev_fields()[3], self.rev_fields()[4], self.rev_fields()[6],
                            self.rev_fields()[8]]
        for item in line_edit_fields:
            for field in item:
                # print(field, field.text())
                if field.text() != '':
                    merge_values.append(field.text())
                else:
                    merge_values.append(None)

        # Need to join the separate street, city, and zip address values into a single address string
        start_index = len(time_boxes + program_boxes) + 5  # index of first address value
        full_address = ', '.join(merge_values[start_index:start_index+3])  # indices of specific lineEdit values
        merge_values[start_index] = full_address
        del merge_values[start_index+1:start_index+3]

        merge_values[len(merge_values) - 2] = merge_values[len(merge_values) - 2][1:]  # strip $ from TotalCost

        # Get values for all textEdit fields and append them to values list
        for item in self.rev_fields()[1]:
            if item.toPlainText() != '':
                merge_values.append(item.toPlainText())
            else:
                merge_values.append(None)

        # Get values for all room comboBoxes
        room_box_values = []
        for item in self.rev_fields()[7]:
            if item.currentText() != '':
                room_box_values.append(item.currentText())
            else:
                room_box_values.append(None)

        # Pull list of merge keys from contract info and zip into dictionary with values list
        merge_keys = self.contract_info()[3]
        # print(len(merge_keys), len(merge_values))
        merge_kp = {k: v for k, v in zip(merge_keys, merge_values)}
        return [merge_kp, room_box_values]

    def browse_rev(self):
        try:
            global path, program_times

            # Open dialog box to select contract file
            root = tkinter.Tk()
            root.withdraw()
            file_path = tkinter.filedialog.askopenfilename(initialdir=path)
            if file_path == '':  # If no file was selected, then return
                return
            else:
                file_path = os.path.abspath(file_path)
            current_folder = file_path.strip(file_path.split('\\')[-1])

            # Check if file selected is a valid contract file
            if self.valid_contract_file_check(file_path) == 1:
                tkinter.messagebox.showerror("Error!", "        Invalid contract file")
                return

            # Display chosen contract file path in GUI
            self.ui.lineEdit_97.setText(os.path.abspath(file_path))

            # Read merge field values from chosen contract file
            try:
                old_kp = self.contract_kp_builder('merge_template.docx', file_path)
            except KeyError:
                tkinter.messagebox.showerror("Error!", "Selected file is not a valid contract")
                return

            # Get original file creation date so it knows where to find the spreadsheet record
            if file_path[-8:-5] == 'REV':
                try:
                    orig_file = [file for file in os.listdir(current_folder) if old_kp["ContractEMS"] in file and
                                 file[-8:-5] not in ['REV', '_CX']]
                    orig_path = os.path.normpath(os.path.join(current_folder, orig_file[0]))
                    # FIXME: temporarily disabled this since incorrect dates have been put on contracts and their
                    # FIXME: records can't be found
                    if "DateWritten" in old_kp and old_kp["DateWritten"] and False:
                        cdate = datetime.datetime.strptime(old_kp["DateWritten"], '%m/%d/%Y').strftime('%b%Y')
                        print("DateWritten timestamp exists")
                    else:
                        cdate = self.creation_date(orig_path)
                        _, found_sheet = self.search_whole_spreadsheet(cdate, old_kp["ContractEMS"])
                        if found_sheet != cdate:
                            cdate = found_sheet
                except Exception as err:
                    tkinter.messagebox.showerror("Error!", str(err) + ' \nCrash log written to error_log.txt')
                    old_stdout = sys.stdout
                    log_file = open('error_log.txt', 'w')
                    sys.stdout = log_file
                    traceback.print_exc()
                    sys.stdout = old_stdout
                    log_file.close()
                    return
            else:
                # FIXME: temporarily disabled this since incorrect dates have been put on contracts and their
                # FIXME: records can't be found
                if "DateWritten" in old_kp and old_kp["DateWritten"] and False:
                    cdate = datetime.datetime.strptime(old_kp["DateWritten"], '%m/%d/%Y').strftime('%b%Y')
                    print("DateWritten timestamp exists")
                else:
                    cdate = self.creation_date(file_path)
                    _, found_sheet = self.search_whole_spreadsheet(cdate, old_kp["ContractEMS"])
                    if found_sheet != cdate:
                        cdate = found_sheet

            # Clear CheckedBy/RevisedBy field when new contract is loaded
            self.ui.lineEdit_99.clear()

            # Read data from excel record to get ContactPhone and room numbers
            xl_record = self.get_xl_record(path+'inhouse_contract_spreadsheet_' + cdate[-4:] + '.xlsx', cdate,
                                           old_kp["ContractEMS"])
            # If record wasn't found, try searching the entire spreadsheet. If this still fails, return.
            if not xl_record:
                xl_record, _ = self.search_whole_spreadsheet(cdate, old_kp["ContractEMS"])
            if not xl_record:
                return

            # Populate contract revision form with values read from chosen contract
            # todo: Clean up
            text_boxes = [self.ui.lineEdit_76, self.ui.lineEdit_77, self.ui.lineEdit_78, self.ui.lineEdit_79,
                          self.ui.lineEdit_80, self.ui.lineEdit_81, self.ui.lineEdit_82, self.ui.lineEdit_110,
                          self.ui.lineEdit_111, self.ui.lineEdit_83,
                          self.ui.lineEdit_84, self.ui.lineEdit_86, self.ui.lineEdit_87, self.ui.lineEdit_88,
                          self.ui.lineEdit_89, self.ui.lineEdit_90, self.ui.textEdit_2, self.ui.textEdit_3,
                          self.ui.lineEdit_34, self.ui.lineEdit_22, self.ui.lineEdit_64, self.ui.lineEdit_61,
                          self.ui.lineEdit_30, self.ui.lineEdit_65, self.ui.lineEdit_69, self.ui.lineEdit_71,
                          self.ui.lineEdit_59, self.ui.lineEdit_56, self.ui.lineEdit_14, self.ui.lineEdit_18,
                          self.ui.lineEdit_60, self.ui.lineEdit_62, self.ui.lineEdit_67, self.ui.lineEdit_68,
                          self.ui.lineEdit_66, self.ui.lineEdit_72, self.ui.lineEdit_46, self.ui.lineEdit_57,
                          self.ui.lineEdit_26, self.ui.lineEdit_42, self.ui.lineEdit_63, self.ui.lineEdit_38,
                          self.ui.lineEdit_74, self.ui.lineEdit_75, self.ui.lineEdit_70, self.ui.lineEdit_73,
                          self.ui.lineEdit_58, self.ui.lineEdit_55, self.ui.lineEdit_85]
            # todo: Clean up
            split_address = [item.strip() for item in str.split(old_kp["SchoolAddress"], ',')]
            kp_values = [old_kp["ContractEMS"], old_kp["BookedBy"], old_kp["ReserverName"], old_kp["ReserverPhone"],
                         old_kp["SchoolName"], old_kp["SchoolFax"], split_address[0], split_address[1],
                         split_address[2], old_kp["SchoolPhone"],
                         old_kp["ContactName"], old_kp["ReserverEmail"], old_kp["DateTaken"], '$' + old_kp["TotalCost"],
                         old_kp["PayDate"], old_kp["DayDate"], old_kp["SpecialConsiderations"], old_kp["RevNote"],
                         old_kp["FirstPgmGrades"], old_kp["SecondPgmGrades"], old_kp["ThirdPgmGrades"],
                         old_kp["FourthPgmGrades"], old_kp["FifthPgmGrades"], old_kp["SixthPgmGrades"],
                         old_kp["SeventhPgmGrades"], old_kp["EighthPgmGrades"], old_kp["NinthPgmGrades"],
                         old_kp["TenthPgmGrades"], old_kp["FirstPgmStudents"], old_kp["SecondPgmStudents"],
                         old_kp["ThirdPgmStudents"], old_kp["FourthPgmStudents"], old_kp["FifthPgmStudents"],
                         old_kp["SixthPgmStudents"], old_kp["SeventhPgmStudents"], old_kp["EighthPgmStudents"],
                         old_kp["NinthPgmStudents"], old_kp["TenthPgmStudents"], old_kp["FirstPgmPrice"],
                         old_kp["SecondPgmPrice"], old_kp["ThirdPgmPrice"], old_kp["FourthPgmPrice"],
                         old_kp["FifthPgmPrice"], old_kp["SixthPgmPrice"], old_kp["SeventhPgmPrice"],
                         old_kp["EighthPgmPrice"], old_kp["NinthPgmPrice"], old_kp["TenthPgmPrice"],
                         str(xl_record[0][9])]

            for index, price in enumerate(kp_values[36:-1]):  # Strip dollar sign and decimal places from prices
                if price is not None:
                    kp_values[index + 36] = price[1:-3]

            fill_kp = {k: v for k, v in zip(text_boxes, kp_values)}
            for key, value in fill_kp.items():
                key.setText(value)

            pgmdate = QtCore.QDate.fromString(old_kp["PgmDate"], 'M/d/yyyy')
            self.ui.calendarWidget_2.setSelectedDate(pgmdate)

            # Populate program times
            pgm_times = [old_kp[self.contract_info()[3][i]] for i in range(10)]

            # avail_times = [None, '10:00-11:00', '11:10-12:10', '12:30-1:30', '1:40-2:40', 'Custom']
            avail_times = [None] + program_times[1:] + ['Custom']
            time_index = []
            for time in pgm_times:
                if time not in avail_times:
                    time_index.append(9)  # set 9 as flag index to indicate custom time
                else:
                    for index, atime in enumerate(avail_times):
                        if time == atime:
                            time_index.append(index)

            time_boxes = self.rev_fields()[2]
            for i in range(len(time_boxes)):
                time_boxes[i].setEditable(False)
                # print('Current time index:', time_index[i])
                if time_index[i] != 9:
                    time_boxes[i].setCurrentIndex(time_index[i])
                else:
                    time_boxes[i].setEditable(True)
                    time_boxes[i].setEditText(pgm_times[i])

            # Populate program titles
            pgm_titles = [old_kp[self.contract_info()[3][i]] for i in range(10, 20)]

            avail_pgms = [None] + self.contract_info()[0]

            pgm_index = []
            custom_pgms = []
            for pgm in pgm_titles:
                if pgm not in avail_pgms:
                    pgm_index.append(len(avail_pgms) - 1)
                    custom_pgms.append(pgm)
                else:
                    for index, apgm in enumerate(avail_pgms):
                        if pgm == apgm:
                            pgm_index.append(index)

            global rev_program_indices  # Globally define indices to check when rev_box_update updates room choices
            rev_program_indices = pgm_index

            # todo: Clean up
            pgm_boxes = [self.ui.comboBox_41, self.ui.comboBox_40, self.ui.comboBox_46, self.ui.comboBox_43,
                         self.ui.comboBox_50, self.ui.comboBox_60, self.ui.comboBox_54, self.ui.comboBox_58,
                         self.ui.comboBox_35, self.ui.comboBox_33]

            # Handle custom programs
            cp_counter = 0
            for i in range(len(pgm_boxes)):
                pgm_boxes[i].setCurrentIndex(pgm_index[i])
                if pgm_index[i] == len(avail_pgms) - 1:
                    pgm_boxes[i].setEditable(True)
                    pgm_boxes[i].setEditText(custom_pgms[cp_counter])
                    cp_counter += 1

            # Populate room number combo boxes with correct room choices for each program loaded from contract, and
            # then set each combo box to room listed in contract
            programs_and_rooms = {k: v for k, v in zip(self.contract_info()[0], self.contract_info()[6])}

            # Make dictionary pairing up program UI boxes and room UI boxes
            program_and_room_boxes = {k: v for k, v in zip(pgm_boxes, self.rev_fields()[7])}

            # Get specific room numbers booked then zip into a dictionary with room combo boxes
            booked_rooms = [xl_record[i][19] for i in range(len(xl_record))]
            rooms_and_booked_kp = {k: v for k, v in zip(self.rev_fields()[7], booked_rooms)}
            # print(booked_rooms, rooms_and_booked_kp)
            # Populate each room UI box with the correct room choices for the program loaded from contract
            for key, value in program_and_room_boxes.items():
                if key.currentText() == '':
                    value.clear()
                elif key.currentText() not in programs_and_rooms:  # This is case when custom program was entered
                    key.setEditable(True)
                    value.setEditable(True)
                    value.setEditText(rooms_and_booked_kp[value])
                elif key.currentText() in programs_and_rooms:
                    value.clear()
                    value.addItems(programs_and_rooms[key.currentText()])
                    # print(key, value)
                    # Check if booked room is partial match for any of room box's choices
                    if any(item for item in programs_and_rooms[key.currentText()] if
                           item.startswith(rooms_and_booked_kp[value])):
                        box_choice = [item for item in programs_and_rooms[key.currentText()] if
                                      item.startswith(rooms_and_booked_kp[value])]
                        # Find index of matching box choice and then change box to that index
                        for index, choice in enumerate(programs_and_rooms[key.currentText()]):
                            if box_choice[0] == choice:
                                value.setCurrentIndex(index)
        except Exception as err:
            tkinter.messagebox.showerror("Error!", str(err) + ' \nCrash log written to error_log.txt')
            old_stdout = sys.stdout
            log_file = open('error_log.txt', 'w')
            sys.stdout = log_file
            traceback.print_exc()
            sys.stdout = old_stdout
            log_file.close()
            return

    def staff_section_file_open(self):
        global path
        root = tkinter.Tk()
        root.withdraw()

        file_path = tkinter.filedialog.askopenfilename(initialdir=path)

        # Check if selected file is valid contract file (if contract file has been selected)
        if file_path != '' and self.valid_contract_file_check(file_path) == 1:
            tkinter.messagebox.showerror("Error!", "Invalid contract file")
            return

        # If file dialog is closed without selecting new file when file path has already been chosen, keep file path
        # that was already entered, otherwise set lineEdit to new file path
        if self.ui.lineEdit_98.text() != '' and file_path == '':
            pass
        else:
            self.ui.lineEdit_98.setText(os.path.abspath(file_path))

        # Section to populate payment info tab fields with any payment info that already exists on contract
        contract_kp = self.contract_kp_builder('merge_template.docx', file_path)
        payment_tab_keys = ["PONumber", "POAmount", "PORecDate", "CheckCCRecDate1", "CheckCCAmount1", "CheckCCRec1",
                            "CheckCCRecDate2", "CheckCCAmount2", "CheckCCRec2", "CheckCCRecDate3", "CheckCCAmount3",
                            "CheckCCRec3", "CheckCC1", "CheckCC2", "CheckCC3", "PaymentInfoComments"]
        # radio buttons are indices 12-17, and textEdit is index 18
        payment_tab_fields = [self.ui.lineEdit_91, self.ui.lineEdit_92, self.ui.lineEdit_93, self.ui.lineEdit_94,
                              self.ui.lineEdit_95, self.ui.lineEdit_96, self.ui.lineEdit_101, self.ui.lineEdit_102,
                              self.ui.lineEdit_103, self.ui.lineEdit_104, self.ui.lineEdit_105, self.ui.lineEdit_106,
                              self.ui.radioButton, self.ui.radioButton_2, self.ui.radioButton_7, self.ui.radioButton_8,
                              self.ui.radioButton_9, self.ui.radioButton_10, self.ui.textEdit_4]
        for i in range(len(payment_tab_keys)):
            if i in range(12):  # lineEdit fields
                if contract_kp[payment_tab_keys[i]] is not None:
                    payment_tab_fields[i].setText(contract_kp[payment_tab_keys[i]].strip('$'))
            elif i in range(12, 15):  # radioButtons
                if contract_kp[payment_tab_keys[i]] == 'CHECK and CC':
                    payment_tab_fields[i-(12-i)].setChecked(True)
                    payment_tab_fields[i+1-(12-i)].setChecked(True)
                elif contract_kp[payment_tab_keys[i]] == 'CHECK':
                    payment_tab_fields[i-(12-i)].setChecked(True)
                elif contract_kp[payment_tab_keys[i]] == 'CC':
                    payment_tab_fields[i+1-(12-i)].setChecked(True)
            elif i == 15:  # textEdit field
                if contract_kp[payment_tab_keys[i]] is not None:
                    payment_tab_fields[i+3].setText(str.split(contract_kp[payment_tab_keys[i]], ': ')[1])

        return

    @staticmethod
    def get_credentials():
        username, password, _ = EntryWidget(2).ok_button()
        if not os.path.exists('credentials'):
            os.mkdir('credentials')
        if not os.path.exists('credentials/user_credentials.dat'):
            open('credentials/user_credentials.dat', 'w+').close()
        if not username and not password:
            return None, None
        # Read any existing usernames and encrypted bConnected keys
        with open('credentials/user_credentials.dat', 'r') as fp:
            cred_list = fp.readlines()
        users = [item.strip('\n').split(',')[0] for item in cred_list]
        # Create instance of class that will encrypt and decrypt user keys
        cipher = CryptoCipher(password)
        # Get new user's bConnected key, encrypt it, then write to credentials file
        if username not in users:
            _, _, user_key_plain = EntryWidget(1).ok_button()
            user_key_enc = cipher.encrypt(user_key_plain)
            with open('credentials/user_credentials.dat', 'a') as fp:
                fp.write(username+','+user_key_enc+'\n')
        # Decrypt existing user's bConnected key
        else:
            user_key_enc = [item.strip('\n').split(',')[1] for item in cred_list if
                            item.strip('\n').split(',')[0] == username][0]
            user_key_plain = cipher.decrypt(user_key_enc)

        return username, user_key_plain

    # todo: THIS SECTION STILL NEEDS WORK
    def staff_section_file_write(self):
        try:
            global path

            file_path = self.ui.lineEdit_98.text()
            current_folder = file_path.strip(file_path.split('\\')[-1])

            root = tkinter.Tk()
            root.withdraw()
            print('Checking if contract file is valid...')
            if self.valid_contract_file_check(file_path) == 1:
                tkinter.messagebox.showerror("Error!", "        Invalid contract file")
                return
            print('Done. Contract file OK.')

            # radio buttons are indices 12-17, and textEdit is index 18
            payment_tab_fields = [self.ui.lineEdit_91, self.ui.lineEdit_92, self.ui.lineEdit_93, self.ui.lineEdit_94,
                                  self.ui.lineEdit_95, self.ui.lineEdit_96, self.ui.lineEdit_101, self.ui.lineEdit_102,
                                  self.ui.lineEdit_103, self.ui.lineEdit_104, self.ui.lineEdit_105, self.ui.lineEdit_106,
                                  self.ui.radioButton, self.ui.radioButton_2, self.ui.radioButton_7, self.ui.radioButton_8,
                                  self.ui.radioButton_9, self.ui.radioButton_10, self.ui.textEdit_4]
            payment_tab_keys = ["PONumber", "POAmount", "PORecDate", "CheckCCRecDate1", "CheckCCAmount1", "CheckCCRec1",
                                "CheckCCRecDate2", "CheckCCAmount2", "CheckCCRec2", "CheckCCRecDate3", "CheckCCAmount3",
                                "CheckCCRec3", "CheckCC1", "CheckCC2", "CheckCC3", "PaymentInfoComments"]

            payment_tab_values = []
            check_flag, cc_flag = 0, 0  # initialize flags for reading radio buttons (check or cc)
            for i in range(len(payment_tab_fields)):
                if i in range(12):  # these are the lineEdit fields
                    if payment_tab_fields[i].text() != '':
                        if i not in [1, 4, 7, 10]:  # indices 1,4,7,10 are payment amounts
                            payment_tab_values.append(payment_tab_fields[i].text())
                        else:
                            payment_tab_values.append('$'+payment_tab_fields[i].text())
                    else:
                        payment_tab_values.append(None)
                elif i in range(12, 18):  # these are the radio buttons
                    if payment_tab_fields[i].isChecked() is True:
                        if i % 2 == 0:  # even indices are for checks
                            check_flag = 1
                            # payment_tab_values.append('CHECK')
                        else:  # odd indices are for cc's
                            cc_flag = 1
                            # payment_tab_values.append('CC')
                    if i % 2 != 0:  # at every odd index, read flags and append appropriate value to dictionary
                        if [check_flag, cc_flag] == [1, 1]:
                            payment_tab_values.append('CHECK and CC')
                        elif check_flag == 1:
                            payment_tab_values.append('CHECK')
                        elif cc_flag == 1:
                            payment_tab_values.append('CC')
                        else:
                            payment_tab_values.append(None)
                        check_flag, cc_flag = 0, 0  # reset flags for next pair of radio buttons
                else:  # this is the textEdit field
                    if payment_tab_fields[i].toPlainText() != '':
                        payment_tab_values.append('PAYMENT INFO COMMENTS: '+payment_tab_fields[i].toPlainText())
                    else:
                        payment_tab_values.append(None)
            # print(payment_tab_values[-4:])
            payment_info = {k: v for k, v in zip(payment_tab_keys, payment_tab_values)}

            # This builds a dictionary using all the merge fields in the contract template and all the merge field
            #  values in the currently selected contract, then updates it with fields/values from the payment info tab
            new_kp = self.contract_kp_builder('merge_template.docx', file_path)
            old_kp = new_kp.copy()  # use for comparison to see which payment fields are new with this update

            for key in payment_info:
                if key in new_kp:  # if merge field exists in dictionary, update value
                    new_kp[key] = payment_info[key]
                else:  # if merge field does not exist in dictionary, add it and its value
                    new_kp.update({key: payment_info[key]})

            ok = tkinter.messagebox.askokcancel("Warning!", "This will overwrite the contract file.\n"
                                                            "Click OK to confirm.")
            if ok != 1:
                tkinter.messagebox.showinfo("Aborted", "Contract update aborted")
                return

            # Get date contract was originally written
            if file_path[-8:-5] == 'REV':
                orig_file = [file for file in os.listdir(current_folder) if new_kp["ContractEMS"] in file and
                             file[-8:-5] not in ['REV', '_CX']]
                orig_path = os.path.join(current_folder, orig_file[0])

                # FIXME: DateWritten timestamp may not be feasible since moving files resets modified date
                # if "DateWritten" not in new_kp or not new_kp["DateWritten"]:
                #     dw = datetime.datetime.fromtimestamp(os.path.getmtime(orig_path)).strftime('%m/%d/%Y')
                #     new_kp.update({"DateWritten": dw})
                c_date = self.creation_date(orig_path)
                _, found_sheet = self.search_whole_spreadsheet(c_date, new_kp["ContractEMS"])
                if found_sheet != c_date:
                    c_date = found_sheet
                # Check if DateWritten timestamp exists in contract. If so, and if it points to the correct spreadsheet
                # page, then pass through the timestamp to the merging dict
                # FIXME
                try:
                    if "DateWritten" in old_kp and old_kp["DateWritten"]:
                        date_for_sheet = datetime.datetime.strptime(old_kp["DateWritten"], '%m/%d/%Y').strftime('%b%Y')
                        if date_for_sheet == found_sheet:
                            new_kp.update({"DateWritten": old_kp["DateWritten"]})
                except:
                    pass
            else:
                # if "DateWritten" not in new_kp or not new_kp["DateWritten"]:
                #     dw = datetime.datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%m/%d/%Y')
                #     new_kp.update({"DateWritten": dw})
                c_date = self.creation_date(file_path)
                _, found_sheet = self.search_whole_spreadsheet(c_date, new_kp["ContractEMS"])
                if found_sheet != c_date:
                    c_date = found_sheet
                # Check if DateWritten timestamp exists in contract. If so, and if it points to the correct spreadsheet
                # page, then pass through the timestamp to the merging dict
                # FIXME
                try:
                    if "DateWritten" in old_kp and old_kp["DateWritten"]:
                        date_for_sheet = datetime.datetime.strptime(old_kp["DateWritten"], '%m/%d/%Y').strftime('%b%Y')
                        if date_for_sheet == found_sheet:
                            new_kp.update({"DateWritten": old_kp["DateWritten"]})
                except:
                    pass

            print('Finished getting file creation date.')

            # Update excel record
            xl_record = self.get_xl_record(path+'inhouse_contract_spreadsheet_'+c_date[-4:]+'.xlsx', c_date,
                                           new_kp["ContractEMS"])

            print("Finished retrieving spreadsheet record.")

            new_kp.update({"ContactPhone": xl_record[0][9]})
            print('Finished updating ContactPhone')
            rev_lines = ["RevLine1", "RevLine2", "RevLine3"]
            for i in range(3):
                print(new_kp[rev_lines[i]])
                if new_kp[rev_lines[i]] is None:
                    if i == 0:
                        break  # if first RevLine is blank, nothing needs to be updated
                    new_kp.update({"RevBy":
                                   new_kp[rev_lines[i-1]][rev_lines[i-1].find('By')+3:rev_lines[i-1].find('By')+13].strip()})
                    print('Found blank revision line and updated dictionary.')
                    break
                elif i == 2:
                    new_kp.update({"RevBy": new_kp[rev_lines[i]][rev_lines[i].find('By')+3:rev_lines[i].find('By')+13].strip()})
            room_numbers = [xl_record[i][19] for i in range(len(xl_record))]
            if xl_record[0][24] is None:
                rev_flag1 = 'not_rev'
                rev_flag2 = 'no'
                print('Found rev/cx cell in record was blank, set revision flags')
            else:
                rev_flag1 = 'rev'
                rev_flag2 = 'yes'
                print('Found rev/cx cell in record was not blank, set revision flags.')

            new_xl_record = self.build_xl_record(new_kp, room_numbers, rev_flag1)
            print('Finished building new excel record.')

            # Write excel record
            try:
                self.write_xl_record(path+'inhouse_contract_spreadsheet_'+c_date[-4:]+'.xlsx', c_date,
                                     new_kp["ContractEMS"], new_xl_record, flag=rev_flag2)
            except PermissionError as err:
                tkinter.messagebox.showerror("Error!", "Permission error. Make sure spreadsheet is not open anywhere, "
                                                       "then try again.\n\n" + str(err))
                return
            print('Finished writing new excel record.')

            # Merge word doc
            out_path = file_path  # adding payment info will always overwrite the current working version
            docxmerge('merge_template.docx', new_kp, out_path)
            print('Finished merging document.')
            tkinter.messagebox.showinfo("Success!", "Contract updated successfully")

            # Find which payment entries are new
            comparison_keys = ["PONumber", "CheckCC1", "CheckCC2", "CheckCC3"]
            new_payment_entries = [item for item in comparison_keys if new_kp[item] and not old_kp[item]]

            # Email payment confirmation
            if new_payment_entries and \
                tkinter.messagebox.askyesno("Email Confirmation", "Would you like to email payment confirmation now?" +
                                            "\n\nClient email: "+new_kp["ReserverEmail"]):
                signed = tkinter.messagebox.askyesno(message="Has a signed copy of the contract been received?")

                subject_string, body_string = self.email_contract_strings(new_kp, 'payment', signed=signed,
                                                                          new_payment_entries=new_payment_entries)
                username, user_key_plain = self.get_credentials()
                if not username and not user_key_plain:
                    tkinter.messagebox.showinfo("Aborted", "Email authentication cancelled, email will not "
                                                           "be sent.\n\nPayment info process is now complete.")
                    return

                send_mail('lhsreg@berkeley.edu', [new_kp["ReserverEmail"], 'lhsreg@berkeley.edu'],
                          subject_string, body_string, files=[], username=username, password=user_key_plain)

                tkinter.messagebox.showinfo("Success!", "Email sent successfully")

            return
        except Exception as err:
            tkinter.messagebox.showerror("Error!", err)
            old_stdout = sys.stdout
            log_file = open("staff_section_log.txt", "w")
            sys.stdout = log_file
            traceback.print_exc()
            sys.stdout = old_stdout
            log_file.close()
            return

    def check_rev_name_change(self, contract_kp, merge_kp):
        # This will check if any of the fields that make up the file name have been changed. If so, a new filename will
        # be returned, otherwise a flag indicating the old filename should be used will be returned
        file_name_fields = ["ContractEMS", "PgmDate", "SchoolName", "PgmDate"]
        file_name_check = [name for name in file_name_fields if contract_kp[name] != merge_kp[name]]

        if not file_name_check:  # empty list is false, so this would mean all the field values match
            return 0
        else:  # if any of the field values don't match, return a new filename with data from the GUI
            file_date = self.ui.calendarWidget_2.selectedDate().toString("MM-dd-yy")
            return file_date + ' ' + str(merge_kp["SchoolName"]) + ' ' + merge_kp["ContractEMS"] + '.docx'

    @staticmethod
    def email_contract_strings(kp, section, signed=None, new_payment_entries=None):
        if not section == 'payment':
            subject_string = "The Lawrence Hall of Science Reservation Contract for " + \
                                                 kp["SchoolName"] + " (" + kp["ContractEMS"] + ")"
            body_string = "Hello,<br><br>" \
                          "Thank you for your reservation! Attached is a PDF copy of your reservation" \
                          " contract & information packet for your visit on <b>"+kp["PgmDate"] + \
                          "</b>, contract number "+kp["ContractEMS"]+".<br><br><font color='red'><b>Full payment is due " \
                                                                     "by "+kp["PayDate"] + \
                          "</b></font>, or your program(s) will be canceled. Payment may be made by purchase order, " \
                          "credit card over" \
                          " the phone (Visa, Mastercard, Discover, or American Express), or a single check made payable " \
                          "to UC Regents. <font color='red'><b>Please submit a signed copy of the contract with your " \
                          "payment.</b></font><br><br>Your reservation will be complete after we receive your payment and " \
                          "signed contract.<br><br>" \
                          "<b>If you have booked a group visit, your payment is due at the door. Please submit a signed " \
                          "copy of the contract to complete your group visit reservation.</b><br><br>" \
                          "<font color='red'><b>Please carefully <u>review the " \
                          "attached contract and information pages</u> for all pertinent " \
                          "information regarding your visit.</b></font><br><br>We require 1 chaperone (admitted free)" \
                          " per 7 students. Additional adults over this number will be charged $12-$14, depending on " \
                          "the program(s) selected.<br><br>Upon arrival please be ready to " \
                          "provide the Visitor Services Desk with a head count of the total number of students and the " \
                          "total number of adult chaperones attending.<br><br>Thank you again for supporting the Lawrence " \
                          "Hall of Science! Visit our " \
                          "<a href='http://www.lawrencehallofscience.org/visit/field_trips#fieldtrips_faq'>FAQ</a> " \
                          "for answers to frequently asked questions.<br><br><br>" + \
                          "Sincerely,<br><br>LHS Registration Staff<br><br><br><br><br>--<br><b>REPLY TO:</b> " + \
                          "lhsreg@berkeley.edu<br><br>Registration<br>The Lawrence Hall of Science<br>1 Centennial Drive #5200<br>" \
                          "Berkeley, CA 94720-5200<br><br>Registration Phone Hours<br>Monday-Friday 8:30 a.m - 4:30 p.m.<br>" \
                          "Registration In-person Hours<br>Wednesday-Friday 8:30 a.m. - 4:30 p.m.<br>Ph: 510-642-5134<br>" \
                          "Fax: 510-643-0994<br>lhsreg@berkeley.edu<br><br>lawrencehallofscience.org<br>"
        else:
            subject_string = "The Lawrence Hall of Science - Payment Confirmation for "+kp["SchoolName"]+" (" + \
                             kp["ContractEMS"]+")"
            body_0 = "Hello,<br><br>"
            body_1 = []
            if "PONumber" in new_payment_entries:
                body_1.append("Thank you for submitting your purchase order for contract "+kp["ContractEMS"] +
                              ". Your purchase order number is <b>"+kp["PONumber"]+"</b> in the amount of <b>" +
                              "$"+str('{0:.2f}'.format(int(kp["POAmount"].strip('$')))) +
                              "</b>. Your group will be invoiced after your visit.<br><br>")
            for item in ["CheckCC1", "CheckCC2", "CheckCC3"]:
                if item in new_payment_entries:
                    body_1.append("Thank you for submitting your payment in the amount of <b>" +
                                  "$"+str('{0:.2f}'.format(int(kp["CheckCCAmount"+str(item[-1])].strip('$')))) +
                                  "</b> for contract "+kp["ContractEMS"] +
                                  ". Your receipt number is <b>"+kp["CheckCCRec"+str(item[-1])]+"</b>.<br><br>")
            body_1 = ''.join(body_1)
            if signed:
                body_2 = 'We have also received your signed contract, so you are all set for your visit.<br><br>'
            else:
                body_2 = 'We have <b>not</b> received your signed contract. Please return a signed copy of your ' \
                         'contract to complete your reservation.<br><br>'
            body_3 = "Thank you for supporting the Lawrence Hall of Science! We look forward to your trip.<br><br>" \
                     "Sincerely,<br>Registration Staff<br><br><br>--<br><b>REPLY TO:</b> lhsreg@berkeley.edu<br><br>" \
                     "Registration<br>The Lawrence Hall of Science<br>1 Centennial Drive #5200<br>Berkeley, CA " \
                     "94720-5200<br><br>Registration Phone Hours<br>Monday-Friday 8:30 a.m. - 4:30 p.m.<br>" \
                     "Registration In-person Hours<br>Wednesday-Friday 8:30 a.m. - 4:30 p.m.<br>Ph: 510-642-5134<br>" \
                     "Fax: 510-643-0994<br>lhsreg@berkeley.edu<br><br>lawrencehallofscience.org"
            body_string = body_0 + body_1 + body_2 + body_3

        return subject_string, body_string

    def check_contract(self):
        try:
            global path
            root = tkinter.Tk()
            root.withdraw()
            file_path = os.path.abspath(self.ui.lineEdit_97.text())
            current_folder = file_path.strip(file_path.split('\\')[-1])
            rev_flag = 'no'  # initialize to non-revised contract state for excel record writing

            # Check to see if revised version of contract exists
            if os.path.exists(file_path[:-5] + '_REV.docx'):
                tkinter.messagebox.showerror("Error!", "Revised version of this contract exists! \n" +
                                                       "Checking must be done on most recent contract version.")
                return

            merge_kp, room_numbers = self.rev_field_scraper()  # Revision tab fields
            checked_date = datetime.datetime.now().strftime("%#m/%#d/%Y")

            # Get current contract field values to check if CheckedBy or RevLine(s) are already filled
            contract_kp = self.contract_kp_builder('merge_template.docx', file_path)  # look in here for existing values
            # Make sure CheckedBy and all RevLines get transferred to merging dict
            merge_kp.update({"PgmDate": str.split(merge_kp["DayDate"], ',')[1].strip()})
            merge_kp.update({"CheckedBy": contract_kp["CheckedBy"], "RevLine1": contract_kp["RevLine1"],
                             "RevLine2": contract_kp["RevLine2"], "RevLine3": contract_kp["RevLine3"]})
            merge_kp.update({"RevDate1": contract_kp["RevDate1"], "RevDate2": contract_kp["RevDate2"],
                             "RevDate3": contract_kp["RevDate3"]})
            payment_info_dict = {k: v for k, v in zip(self.contract_info()[5],
                                                      [contract_kp[item] for item in self.contract_info()[5]])}
            merge_kp.update(payment_info_dict)
            checked_keys = ["CheckedBy", "RevLine1", "RevLine2", "RevLine3"]
            checked_by_values = [contract_kp[item] for item in checked_keys]  # this preserves order

            # Get program date for sorting contract file after checking. Format will be mmmYYYY, e.g. Dec2016
            booked_date = datetime.datetime.strptime(merge_kp["PgmDate"], '%m/%d/%Y').strftime('%b%Y')

            # Find the first empty initials field and fill it with new initials. This must also work for checking revised
            # contracts, so the RevLine must be retrieved, stripped into separate parts, and then rebuilt with the new
            # initials in the final "By" field.
            for i in range(len(checked_by_values)):
                print(checked_by_values[i])
                if checked_by_values[i] is None or i == 3:
                    if i in [0, 1]:  # contract either not checked, or checked but not revised
                        if i == 1:
                            recheck = tkinter.messagebox.askyesno("Re-check?", "This contract has "
                                                                               "already been checked.\n"
                                                                               "Would you like to re-check it?")
                            if recheck != 1:
                                tkinter.messagebox.showinfo("Aborted", "Contract update aborted")
                                return

                        merge_kp.update({"CheckedBy": merge_kp["RevBy"]})  # uses same lineEdit as revision
                        merge_kp["RevBy"] = None

                    else:  # Case where contract has been checked and at least one RevLine exists
                        rev_flag = 'yes'  # contract has been revised, set rev_flag for excel writing
                        if i < 3 or checked_by_values[i] is None:
                            working_rev_line = checked_by_values[i-1]  # need the RevLine just BEFORE the first blank one
                        else:
                            working_rev_line = checked_by_values[i]
                        initials_check = working_rev_line[-2:].strip()  # trying to grab checkedby initials from RevLine
                        # print(initials_check)
                        if not initials_check:
                            pass
                        else:
                            recheck = tkinter.messagebox.askyesno("Re-check?", "This revision of the contract has "
                                                                               "already been checked.\n"
                                                                               "Would you like to re-check it?")
                            if recheck != 1:
                                tkinter.messagebox.showinfo("Aborted", "Contract update aborted")
                                return

                        rev_date = working_rev_line[8:28].strip()  # does this grab the date well enough?
                        rev_by_index = working_rev_line.find("By") + 3  # index of location after first "By:"
                        rev_by_index_stop = working_rev_line.find("Reviewed/Sent") - 1
                        revised_by = working_rev_line[rev_by_index:rev_by_index_stop].strip()
                        first_half = 'REVISED:    '+rev_date+'                 By:     '+revised_by+'           '
                        second_half = '          Reviewed/Sent:   '+checked_date+'             '+'By:    '+merge_kp["RevBy"]
                        # RevBy field is same as CheckedBy on GUI
                        new_rev_line = first_half+second_half
                        print(new_rev_line)
                        if i < 3 or checked_by_values[i] is None:
                            merge_kp.update({checked_keys[i-1]: new_rev_line})
                        else:
                            merge_kp.update({checked_keys[i]: new_rev_line})
                    break  # stop after finding first entry to fill

            # If existing contract has not been checked then use today's date as checked by date, otherwise re-merge
            # the stored checked by date
            if not contract_kp["CheckedDate"]:
                merge_kp.update({"CheckedDate": checked_date})
            else:
                merge_kp.update({"CheckedDate": contract_kp["CheckedDate"]})

            # Format prices with $ and decimal places before merging to word doc
            for item in self.contract_info()[3][54:64]:
                if merge_kp[item] is not None:
                    merge_kp[item] = '$' + str('{:.2f}'.format(int(merge_kp[item])))

            # Check to see if no contract has been loaded
            if file_path == '':
                tkinter.messagebox.showerror("Error", "     Please select a contract file.")
                return

            # Check for errors before writing data
            if self.contract_field_error_check('rev', 'no') == 1:
                print('Error detected. Halting execution.')
                return

            # Get old excel record to check length
            # Find original date contract was written
            rev_flag2 = 'not_rev'  # initialize to non-revised state for building new excel record
            if file_path[-8:-5] == 'REV':
                rev_flag2 = 'rev'

                # This will grab portion of existing file path that should match any revised or canceled versions
                file_tail = os.path.split(file_path)[1][:-9]
                orig_file = [file for file in os.listdir(current_folder) if contract_kp["ContractEMS"] in file and
                             file[-8:-5] not in ['REV', '_CX']]
                if not orig_file:  # if list returns empty, give error and return
                    tkinter.messagebox.showerror("Error!", "No un-revised version of this contract was found")
                    return
                else:  # if list returns file, build file path for sending to creation_date module
                    orig_path = os.path.normpath(os.path.join(current_folder, orig_file[0]))

                # if "DateWritten" not in contract_kp or not contract_kp["DateWritten"]:
                #     dw = datetime.datetime.fromtimestamp(os.path.getmtime(orig_path)).strftime('%m/%d/%Y')
                #     merge_kp.update({"DateWritten": dw})
                c_date = self.creation_date(orig_path)
                _, found_sheet = self.search_whole_spreadsheet(c_date, contract_kp["ContractEMS"])
                if found_sheet != c_date:
                    c_date = found_sheet
                # Check if DateWritten timestamp exists in contract. If so, and if it points to the correct spreadsheet
                # page, then pass through the timestamp to the merging dict
                # FIXME
                try:
                    if "DateWritten" in contract_kp and contract_kp["DateWritten"]:
                        date_for_sheet = datetime.datetime.strptime(contract_kp["DateWritten"], '%m/%d/%Y').strftime('%b%Y')
                        if date_for_sheet == found_sheet:
                            merge_kp.update({"DateWritten": contract_kp["DateWritten"]})
                except:
                    pass
            else:

                # if "DateWritten" not in contract_kp or not contract_kp["DateWritten"]:
                #     dw = datetime.datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%m/%d/%Y')
                #     merge_kp.update({"DateWritten": dw})
                c_date = self.creation_date(file_path)
                _, found_sheet = self.search_whole_spreadsheet(c_date, contract_kp["ContractEMS"])
                if found_sheet != c_date:
                    c_date = found_sheet

                # Check if DateWritten timestamp exists in contract. If so, and if it points to the correct spreadsheet
                # page, then pass through the timestamp to the merging dict
                # FIXME
                try:
                    if "DateWritten" in contract_kp and contract_kp["DateWritten"]:
                        date_for_sheet = datetime.datetime.strptime(contract_kp["DateWritten"], '%m/%d/%Y').strftime('%b%Y')
                        if date_for_sheet == found_sheet:
                            merge_kp.update({"DateWritten": contract_kp["DateWritten"]})
                except:
                    pass

            xl_record = self.get_xl_record(path+'inhouse_contract_spreadsheet_' + c_date[-4:] + '.xlsx', c_date,
                                           contract_kp["ContractEMS"])
            if xl_record is None:  # this means get_xl_record had an error
                return

            # Build new excel record
            if merge_kp["RevBy"] is None and rev_flag2 == 'rev':
                merge_kp.update({"RevBy": ''})
            new_xl_record = self.build_xl_record(merge_kp, room_numbers, rev_flag2)
            if new_xl_record is None:  # this means build_xl_record had an error
                return

            # Make sure new record is same length as old (same number of programs)
            if len(xl_record) != len(new_xl_record[0]):  # columns and rows are switched between these two
                tkinter.messagebox.showerror("Error!",
                                             "Changing number of programs must be done as a revision!\n"
                                             "This contract must have " + str(len(xl_record)) +
                                             " workshops booked.")
                return

            # Determine and create directory for contract to be filed into after checking
            months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            for i, m in enumerate(months):
                if booked_date[:3] == m:
                    if i % 2 == 0:
                        out_folder = 'Inhouse '+months[i]+'-'+months[i+1]+' '+booked_date[3:]
                    else:
                        out_folder = 'Inhouse '+months[i-1]+'-'+months[i]+' '+booked_date[3:]
                    out_dir = os.path.abspath(os.path.join(path, out_folder))
                    if not os.path.exists(out_dir):
                        os.mkdir(out_dir)
            print(out_dir)
            # Check if this is will be a re-filing case. If so, find related files to be moved
            other_files = None
            if "Inhouse " in current_folder and current_folder != out_dir:
                other_files = [file for file in os.listdir(current_folder) if contract_kp["ContractEMS"] in file]

            # Determine correct filename that should be saved to
            if self.check_rev_name_change(contract_kp, merge_kp) == 0:
                out_path = os.path.join(out_dir, str.split(file_path, '\\')[-1])
            else:
                if file_path[-8:-5] == 'REV':
                    out_path = os.path.join(out_dir, self.check_rev_name_change(contract_kp, merge_kp)[:-5]+'_REV.docx')
                else:
                    out_path = os.path.join(out_dir, self.check_rev_name_change(contract_kp, merge_kp))

            # Confirm decision to update
            ok = tkinter.messagebox.askokcancel("Warning!", "       This will overwrite the contract file. Continue?")
            if ok == 1:
                # Update excel record
                self.write_xl_record(path+'inhouse_contract_spreadsheet_' + c_date[-4:] + '.xlsx', c_date,
                                     contract_kp["ContractEMS"], new_xl_record, flag=rev_flag)

                # If file name has changed, but only by capitalization, delete old file before merging new one to
                # prevent old file name being re-used by case-insensitive file system (Windows)
                if os.path.abspath(file_path.lower()) == os.path.abspath(out_path.lower()):
                    os.remove(file_path)

                # Merge word doc
                docxmerge('merge_template.docx', merge_kp, out_path)

                # Save PDF copy
                self.save_as_pdf(out_path, out_path[:-4]+'pdf')

                # If re-filing, move any related files
                if other_files:
                    for file in other_files:
                        os.rename(os.path.join(current_folder, file), os.path.join(out_dir, file))

                # Email PDF to contact email address
                email_check = tkinter.messagebox.askyesno("Email Contract", "Would you like to email this contract to "
                                                                            "the client now?\n\nContract file: " +
                                                                            out_path[:-4]+'pdf\n\n'+"Client email: " +
                                                                            merge_kp["ReserverEmail"])
                if email_check:
                    cf = Popen([out_path[:-4]+'pdf'], shell=True)
                    if tkinter.messagebox.askokcancel("Confirm Email", "Please review the contract and then click OK "
                                                                       "\nto send the email or Cancel to abort"):

                        try:
                            username, user_key_plain = self.get_credentials()

                            # Find info packet pdf
                            info_packet_file = [file for file in os.listdir('LHS_Full_Information_Packet')][0]
                            info_packet_path = os.path.abspath(os.path.join('LHS_Full_Information_Packet',
                                                                            info_packet_file))

                            # Build strings for email subject and body
                            subject_string, body_string = self.email_contract_strings(merge_kp, 'check')

                            # Send email with pdf contract attached
                            send_mail('lhsreg@berkeley.edu', [merge_kp["ReserverEmail"], 'lhsreg@berkeley.edu'],
                                      subject_string,
                                      body_string,
                                      files=[out_path[:-4]+'pdf', info_packet_path],
                                      username=username,
                                      password=user_key_plain)
                            tkinter.messagebox.showinfo("Email Sent", "Contract sent successfully")

                        except Exception as err:
                            tkinter.messagebox.showerror("Error!", "Email authentication error. Contract "
                                                                   "will not be emailed.\n\n" + str(err))

                    else:
                        tkinter.messagebox.showinfo("Email Aborted", "Contract will not be emailed")
                    cf.kill()
            else:
                tkinter.messagebox.showinfo("Aborted", message="Contract update aborted")
                return

            # Set file path field to new contract file
            self.ui.lineEdit_97.setText(os.path.abspath(out_path))

            # Clear checker/reviser initials to prevent an accidental re-check/revise
            self.ui.lineEdit_99.clear()

            # If file name has changed by more than just capitalization, delete the old file name and old pdf
            if os.path.abspath(file_path.lower()) != os.path.abspath(out_path.lower()) and not other_files:
                os.remove(file_path)
                if os.path.exists(file_path[:-4]+'pdf'):
                    os.remove(file_path[:-4]+'pdf')

            tkinter.messagebox.showinfo("Success!", "Contract checking completed")

            return
        except PermissionError as err:
            tkinter.messagebox.showerror("Error!", "Permission error. Make sure contract file and spreadsheet are not "
                                                   "open anywhere, then try again.\n\n" + str(err))
            return
        except Exception as err:
            tkinter.messagebox.showerror("Error!", str(err) + ' \nCrash log written to error_log.txt')
            old_stdout = sys.stdout
            log_file = open('error_log.txt', "w")
            sys.stdout = log_file
            traceback.print_exc()
            sys.stdout = old_stdout
            log_file.close()
            return

    def apply_rev(self):
        try:
            global path
            root = tkinter.Tk()
            root.withdraw()
            ok = tkinter.messagebox.askokcancel("Continue?", "Click OK to confirm contract revision")
            if ok == 0:
                return

            file_path = self.ui.lineEdit_97.text()  # Loaded contract file
            current_folder = file_path.strip(file_path.split('\\')[-1])

            # Scrape all revision tab field values and return in dictionary
            merge_kp, room_box_values = self.rev_field_scraper()
            merge_kp.update({"PgmDate": str.split(merge_kp["DayDate"], ',')[1].strip()})  # Add program date without day
            for price_field in self.contract_info()[3][54:64]:  # Format program prices for merging
                if merge_kp[price_field] is not None:
                    merge_kp[price_field] = '$' + str('{:.2f}'.format(int(merge_kp[price_field])))

            rev_date = datetime.datetime.now().strftime("%#m/%#d/%Y")

            contract_kp = self.contract_kp_builder('merge_template.docx', file_path)

            # Check if the contract has been checked before doing a revision
            if contract_kp["CheckedBy"] is None:
                tkinter.messagebox.showerror("Error!", "Contract must be checked before an official revision can"
                                                       " be made!")
                return

            # make sure merging dict has CheckedBy and CheckedDate fields
            merge_kp.update({"CheckedBy": contract_kp["CheckedBy"], "CheckedDate": contract_kp["CheckedDate"]})
            payment_info_dict = {k: v for k, v in zip(self.contract_info()[5],
                                                      [contract_kp[item] for item in self.contract_info()[5]])}
            merge_kp.update(payment_info_dict)
            # todo: uncomment after all current contracts have DateWritten field filled
            # merge_kp.update({"DateWritten": contract_kp["DateWritten"]})

            # Read revision dates from contract, fill first blank with new revision date
            rev_date_keys = ["RevDate1", "RevDate2", "RevDate3"]
            rev_date_values = [contract_kp[key] for key in rev_date_keys]  # revision date values from loaded contract
            for i in range(len(rev_date_keys)):  # this method preserves order, dictionary does not
                if rev_date_values[i] is None:
                    if i == 0:
                        merge_kp.update({rev_date_keys[i]: "REVISED:   "+rev_date})  # fill blank revision date field
                    else:
                        merge_kp.update({rev_date_keys[i]: ',   '+rev_date})
                    break  # stop after filling first blank revision date field
                else:
                    merge_kp.update({rev_date_keys[i]: rev_date_values[i]})  # transfer existing values to merging dict

            rev_line_fields = ["RevLine1", "RevLine2", "RevLine3"]
            # todo: Think of better way to make revision line. Maybe switch to a table
            # The info that needs to be included is: REVISED, By, Reviewed/Sent, By.
            # Data are stored in: rev_date, merge_kp["RevBy"], and last two should be blank, filled at next contract check
            new_rev_line = 'This is the placeholder string for the revision line.'
            first_half = 'REVISED:    '+rev_date+'                   By:    '+merge_kp["RevBy"]+'           '
            second_half = '            Reviewed/Sent:    '+'                                '+'  By:    '+'      '
            new_rev_line = first_half + second_half
            # This assigns the current revision field data to the first empty RevLine key of the three possible.
            for field in rev_line_fields:
                merge_kp.update({field: contract_kp[field]})  # transfer existing RevLine values to dict to be merged
                print('Checking', field+'...')
                print(field, contract_kp[field])
                if contract_kp[field] is None:
                    print(field + ' is empty.')
                    merge_kp.update({field: new_rev_line})  # placeholder for rev line created with this revision
                    print(merge_kp[field])
                    break  # stop after filling first None field

            print('Revised by and note:', merge_kp["RevBy"], merge_kp["RevNote"])
            print('CheckedBy:', merge_kp["CheckedBy"])

            # Check to see that a contract has been loaded
            if file_path == '':
                tkinter.messagebox.showerror("Error!", "Please select a contract file.")
                return

            # Determine correct filename to save to
            other_files = None
            if self.check_rev_name_change(contract_kp, merge_kp) == 0:  # case where filename fields haven't changed
                if file_path[-8:-5] != 'REV':
                    out_path = file_path[:-5] + '_REV.docx'  # save revised contract with _REV added to filename
                else:
                    out_path = file_path  # if revised contract file exists, new revisions overwrite it
            else:  # case where filename fields have changed
                print("Filename changed")
                changed_name = self.check_rev_name_change(contract_kp, merge_kp)
                booked_date = datetime.datetime.strptime(changed_name[:8], '%m-%d-%y').strftime('%b%Y')

                # check if contract needs to be re-filed in different folder due to booking date change
                if not booked_date[:3] in current_folder or not booked_date[3:] in current_folder:
                    print("Re-filing contract")
                    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                    for i, m in enumerate(months):
                        if booked_date[:3] == m:
                            if i % 2 == 0:
                                out_folder = 'Inhouse '+months[i]+'-'+months[i+1]+' '+booked_date[3:]
                            else:
                                out_folder = 'Inhouse '+months[i-1]+'-'+months[i]+' '+booked_date[3:]
                            out_dir = os.path.abspath(os.path.join(path, out_folder))
                            if not os.path.exists(out_dir):
                                os.mkdir(out_dir)
                    out_path = os.path.join(out_dir, changed_name[:-5] + '_REV.docx')

                    # find all other files related to this contract
                    other_files = [file for file in os.listdir(current_folder) if contract_kp["ContractEMS"] in file]
                else:
                    out_path = os.path.join(current_folder, changed_name[:-5] + '_REV.docx')
                # file_delete = 1

            # Check for errors before writing data
            if self.contract_field_error_check('rev', 'yes') == 1:
                print('Error detected. Halting execution.')
                return

            # Get contract file creation date in order to find record in spreadsheet
            if file_path[-8:-5] == 'REV':
                orig_file = [file for file in os.listdir(current_folder)if contract_kp["ContractEMS"]
                             in file and file[-8:-5] not in ['REV', '_CX']]
                if not orig_file:
                    tkinter.messagebox.showerror("Error!", "        Un-revised contract file not found.")
                    return
                orig_path = os.path.join(current_folder, orig_file[0])

                # if "DateWritten" not in contract_kp or not contract_kp["DateWritten"]:
                #     dw = datetime.datetime.fromtimestamp(os.path.getmtime(orig_path)).strftime('%m/%d/%Y')
                #     merge_kp.update({"DateWritten": dw})
                c_date = self.creation_date(orig_path)
                _, found_sheet = self.search_whole_spreadsheet(c_date, contract_kp["ContractEMS"])
                if found_sheet != c_date:
                    c_date = found_sheet
                # Check if DateWritten timestamp exists in contract. If so, and if it points to the correct spreadsheet
                # page, then pass through the timestamp to the merging dict
                # FIXME
                try:
                    if "DateWritten" in contract_kp and contract_kp["DateWritten"]:
                        date_for_sheet = datetime.datetime.strptime(contract_kp["DateWritten"], '%m/%d/%Y').strftime('%b%Y')
                        if date_for_sheet == found_sheet:
                            merge_kp.update({"DateWritten": contract_kp["DateWritten"]})
                except:
                    pass
            else:
                # This checks if unrevised contract has been loaded by mistake
                if os.path.exists(file_path[:-5] + '_REV.docx'):
                    tkinter.messagebox.showerror("Error!",
                                                 "A revised version of this contract already exists!\n"
                                                 "Please load the revised contract to make further changes.")
                    return
                else:
                    # if "DateWritten" not in contract_kp or not contract_kp["DateWritten"]:
                    #     dw = datetime.datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%m/%d/%Y')
                    #     merge_kp.update({"DateWritten": dw})
                    c_date = self.creation_date(file_path)
                    _, found_sheet = self.search_whole_spreadsheet(c_date, contract_kp["ContractEMS"])
                    if found_sheet != c_date:
                        c_date = found_sheet
                    # Check if DateWritten timestamp exists in contract. If so, and if it points to the correct
                    # spreadsheet page, then pass through the timestamp to the merging dict
                    # FIXME
                    try:
                        if "DateWritten" in contract_kp and contract_kp["DateWritten"]:
                            date_for_sheet = datetime.datetime.strptime(contract_kp["DateWritten"], '%m/%d/%Y').strftime('%b%Y')
                            if date_for_sheet == found_sheet:
                                merge_kp.update({"DateWritten": contract_kp["DateWritten"]})
                    except:
                        pass

            # Update excel record
            new_xl_record = self.build_xl_record(merge_kp, room_box_values, 'rev')
            self.write_xl_record(path+'inhouse_contract_spreadsheet_' + c_date[-4:] + '.xlsx', c_date,
                                 contract_kp["ContractEMS"], new_xl_record)

            # Merge word doc
            print('Merging data to', out_path)
            docxmerge('merge_template.docx', merge_kp, out_path)

            # If contract was re-filed move any related files to new directory
            if other_files:
                for file in other_files:
                        os.rename(os.path.join(current_folder, file), os.path.join(out_dir, file))

            self.ui.lineEdit_97.setText(os.path.abspath(out_path))
            self.ui.lineEdit_99.clear()
            # if file_delete == 1:  # if saving to a different filename, delete old one
            #     os.remove(file_path)

            tkinter.messagebox.showinfo("Success!", "Contract revision completed successfully")

            return
        except PermissionError as err:
            tkinter.messagebox.showerror("Error!", "Permission error. Make sure the following file is "
                                                   "not open anywhere, then try again.\n\n" + str(err))
            return
        except Exception as err:
            tkinter.messagebox.showerror("Error!", str(err)+'\nCrash log written to error_log.txt')
            old_stdout = sys.stdout
            log_file = open('error_log.txt', 'w')
            sys.stdout = log_file
            traceback.print_exc()
            sys.stdout = old_stdout
            log_file.close()
            return

    def valid_contract_file_check(self, file):
        error_flag = 0
        try:
            test_kp = self.contract_kp_builder('merge_template.docx', file)
            if test_kp["ContractEMS"] is None or len(test_kp) < 10:
                error_flag = 1
        except Exception as err:
            print('Error:', err)
            error_flag = 1

        return error_flag

    def contract_field_error_check(self, tab, *is_rev):
        # This section will check for any errors, incompletions, etc. in the new contract form or contract revision
        # form. It will handle error dialogs and then return a flag to the calling function to halt execution.
        error_flag = 0

        # This gets all the top-area fields for the new contract tab and the revision tab
        new_fields = self.new_contract_fields()[6]
        rev_fields = self.rev_fields()[0][:-1] + [self.rev_fields()[8][1]]
        field_labels = ["Contract #", "Booked By", "Date Taken", "School Name", "School Fax", "School Street Address",
                        "School City", "School ZIP",
                        "School Phone", "Reserver Name", "Reserver Phone", "Reserver Email", "Contact Name",
                        "Contact Phone", "Program Date", "Payment Due Date"]
        new_field_dict = {k: v for k, v in zip(new_fields, field_labels)}
        rev_field_dict = {k: v for k, v in zip(rev_fields, field_labels)}

        messages = []  # this will hold all the error messages
        if tab == 'new':  # this section will check for errors when creating a new contract
            print('New contract tab error checker')

            for field in new_fields:
                if new_field_dict[field] != "School Fax" and field.text() == '':
                    print('Please enter a ' + new_field_dict[field])
                    messages.append('Please enter a ' + new_field_dict[field])
                    error_flag = 1
                # elif new_field_dict[field] == "School Address":
                #     print(len(str.split(field.text(), ',')))
                #     if len(str.split(field.text(), ',')) < 3:
                #         print('Please enter a complete School Address with format: street, city, ZIP')
                #         messages.append('Please enter a complete School Address with format: street, city, ZIP')
                #         error_flag = 1

                program_block = [self.new_contract_fields()[0], self.new_contract_fields()[4],
                                 self.new_contract_fields()[5], self.new_contract_fields()[2],
                                 self.new_contract_fields()[1]]

        elif tab == 'rev':  # this section will check for errors when checking or revising a contract
            print('Check/Revise tab error checker')
            # Here is where to check all fields not in the program row block
            rev_field_dict.update({self.rev_fields()[0][13]: "your initials"})
            for field in rev_field_dict:
                if rev_field_dict[field] != "School Fax" and field.text() == '':
                    if rev_field_dict[field] != "your initials":
                        print('Please enter a ' + rev_field_dict[field])
                        messages.append('Please enter a ' + rev_field_dict[field])
                    else:
                        print('Please enter ' + rev_field_dict[field])
                        messages.append('Please enter ' + rev_field_dict[field])
                    error_flag = 1

            if is_rev[0] == 'yes':  # if update is a revision, check that revision note is entered
                if self.ui.textEdit_3.toPlainText() == '':
                    messages.append('Please enter a revision note')
                    print('Please enter a revision note')
                    error_flag = 1

            program_block = [self.rev_fields()[2], self.rev_fields()[3], self.rev_fields()[4], self.rev_fields()[5],
                             self.rev_fields()[6]]

        # Here is where to check program rows (at least one row filled, no partials, no non-consecutive rows)
        combo_box_columns = [0, 3]
        blank_counter = []
        consecutive_error_flag = 0
        for i in range(10):
            blank_counter.append(0)
            for j in range(len(program_block)):
                if j in combo_box_columns and program_block[j][i].currentText() == '':
                    blank_counter[i] += 1
                elif j not in combo_box_columns and program_block[j][i].text() == '':
                    blank_counter[i] += 1
            if i == 0 and blank_counter[i] == 5:
                # This means the first program row is blank
                print('placeholder for first row blank error')
                messages.append("First program row is blank")
                error_flag = 1
            elif 0 < blank_counter[i] < 5:
                if blank_counter[i-1] == 5:
                    # This means the current row is partially filled and the previous row was blank (non-consecutive)
                    print('placeholder for non-consecutive row error')
                # This means the current row is partially filled
                print('placeholder for incomplete program row error')
                print('row', i+1, 'is incomplete!')
                messages.append("Program row " + str(i+1) + " is incomplete")
                error_flag = 1
            elif i > 0 and blank_counter[i] == 0 and blank_counter[i-1] == 5:
                # This means the current row is not blank but the previous one is (non-consecutive)
                print('placeholder for non-consecutive program row error')
                if consecutive_error_flag != 1:
                    messages.append("Programs must be entered on consecutive rows")
                    consecutive_error_flag = 1
                error_flag = 1

        if error_flag == 1:
            messages = '\n'.join(messages)
            tkinter.messagebox.showerror("Error!", "Please correct the following errors:\n\n" + messages)

        return error_flag

    def time_box_update(self):
        for item in self.new_contract_fields()[0]:
            if item.currentText() in self.contract_info()[4]:
                item.setEditable(False)
            elif item.currentText() == 'Custom':
                item.setEditable(True)
            else:
                item.setEditable(True)

    def rev_time_box_update(self):
        # print('rev_time_box_update was called!')
        for item in self.rev_fields()[2]:
            if item.currentText() in self.contract_info()[4]:
                item.setEditable(False)
            elif item.currentText() == 'Custom':
                item.setEditable(True)
            else:
                item.setEditable(True)

    def show_date(self):  # Get date from calendar and display in lineEdit box
        day_and_date = self.ui.calendarWidget.selectedDate().toString("dddd, M/d/yyyy")
        self.ui.lineEdit_53.setText(day_and_date)

    def rev_show_date(self):
        day_and_date = self.ui.calendarWidget_2.selectedDate().toString("dddd, M/d/yyyy")
        self.ui.lineEdit_90.setText(day_and_date)

    def cost_calc(self):  # Calculate total cost from program prices that have been entered
        prices = self.new_contract_fields()[1]
        total = 0
        for price in prices:
            if price.text() != '':
                total += int(price.text())
        self.ui.lineEdit_50.setText('$' + '{:.2f}'.format(total))
        return '{:.2f}'.format(total)

    def room_list(self):  # Populate room choices when program title is selected
        global new_program_indices
        # program_boxes = self.new_contract_fields()[2]
        program_indices = [box.currentIndex() for box in self.new_contract_fields()[2]]
        # room_boxes = self.new_contract_fields()[3]
        # room_groups = self.contract_info()[1]
        # index_map = self.contract_info()[2]
        # pgm_titles = self.contract_info()[0]
        # pgm_rooms = self.contract_info()[6]
        boxes = {k: v for k, v in zip(self.new_contract_fields()[2], self.new_contract_fields()[3])}
        # program_room_map = {k: v for k, v in zip(program_titles, index_map)}
        programs_and_rooms = {k: v for k, v in zip(self.contract_info()[0], self.contract_info()[6])}
        # Check if global index variable has been assigned, and then update only those rows that have changed
        if new_program_indices is not None:
            programs_and_indices = {k: v for k, v in zip(self.new_contract_fields()[2], new_program_indices)}
            for key, value in boxes.items():
                if key.currentIndex() != programs_and_indices[key]:
                    if key.currentText() == '':
                        key.setEditable(False)
                        value.setEditable(False)
                        value.clear()
                    # elif key.currentText() not in program_room_map:
                    elif key.currentText() not in programs_and_rooms:
                        key.setEditable(True)
                        value.clear()
                        value.setEditable(True)
                    # elif key.currentText() in program_room_map:
                    elif key.currentText() in programs_and_rooms:
                        key.setEditable(False)
                        value.setEditable(False)
                        value.clear()
                        # value.addItems(room_groups[program_room_map[key.currentText()]])
                        value.addItems(programs_and_rooms[key.currentText()])
        else:  # If it's the first run of this function, update all room box choices
            for key, value in boxes.items():
                if key.currentText() == '':
                    key.setEditable(False)
                    value.setEditable(False)
                    value.clear()
                # elif key.currentText() not in program_room_map:
                elif key.currentText() not in programs_and_rooms:
                    key.setEditable(True)
                    value.clear()
                    value.setEditable(True)
                # elif key.currentText() in program_room_map:
                elif key.currentText() in programs_and_rooms:
                    key.setEditable(False)
                    value.setEditable(False)
                    value.clear()
                    # value.addItems(room_groups[program_room_map[key.currentText()]])
                    value.addItems(programs_and_rooms[key.currentText()])

        new_program_indices = program_indices  # Update global variable with new program indices
        return

    def merge(self):  # Write spreadsheet file and merge word doc contract
        global path

        # Section to define all variables
        # FileName='../mm_contract/merge_template.docx'
        FileName = 'merge_template.docx'
        ContractEMS = self.ui.lineEdit_3.text()
        BookedBy = self.ui.lineEdit_52.text()
        DateTaken = self.ui.lineEdit_10.text()
        PgmDate=self.ui.lineEdit_53.text()
        FileDate = self.ui.calendarWidget.selectedDate().toString("MM-dd-yy")
        SchoolName=self.ui.lineEdit_4.text()
        JustDay = self.ui.calendarWidget.selectedDate().toString("dddd")
        JustDate = self.ui.calendarWidget.selectedDate().toString("M/d/yyyy")
        # SchoolAddress = self.ui.lineEdit_5.text()
        school_street = self.ui.lineEdit_5.text()
        school_city = self.ui.lineEdit_108.text()
        school_zip = self.ui.lineEdit_109.text()
        school_address = ', '.join([item.strip() for item in [school_street, school_city, school_zip]])
        # SplitSchoolAddress = str.split(self.ui.lineEdit_5.text(), ',')
        # for index in range(len(SplitSchoolAddress)):
        #     SplitSchoolAddress[index] = SplitSchoolAddress[index].strip()

        SchoolPhone = self.ui.lineEdit_8.text()
        SchoolFax = self.ui.lineEdit_2.text()
        ReserverName = self.ui.lineEdit_6.text()
        ReserverPhone = self.ui.lineEdit_7.text()
        ReserverEmail = self.ui.lineEdit_9.text()
        ContactName = self.ui.lineEdit.text()
        ContactPhone = self.ui.lineEdit_54.text()
        PayDate = self.ui.lineEdit_51.text()
        OutFile = FileDate + ' ' + SchoolName + ' ' + ContractEMS + '.docx'
        outpath = path + OutFile
        FirstPgmTime = self.ui.comboBox_21.currentText()
        FirstPgmGrades = self.ui.lineEdit_11.text()
        FirstPgmStudents = self.ui.lineEdit_12.text()
        FirstPgmPrice = self.ui.lineEdit_13.text()
        SecondPgmTime = self.ui.comboBox_22.currentText()
        SecondPgmGrades = self.ui.lineEdit_15.text()
        SecondPgmStudents = self.ui.lineEdit_16.text()
        SecondPgmPrice = self.ui.lineEdit_17.text()
        ThirdPgmTime = self.ui.comboBox_23.currentText()
        ThirdPgmGrades = self.ui.lineEdit_19.text()
        ThirdPgmStudents = self.ui.lineEdit_20.text()
        ThirdPgmPrice = self.ui.lineEdit_21.text()
        FourthPgmTime = self.ui.comboBox_24.currentText()
        FourthPgmGrades = self.ui.lineEdit_23.text()
        FourthPgmStudents = self.ui.lineEdit_24.text()
        FourthPgmPrice = self.ui.lineEdit_25.text()
        FifthPgmTime = self.ui.comboBox_25.currentText()
        FifthPgmGrades = self.ui.lineEdit_27.text()
        FifthPgmStudents = self.ui.lineEdit_28.text()
        FifthPgmPrice = self.ui.lineEdit_29.text()
        SixthPgmTime = self.ui.comboBox_26.currentText()
        SixthPgmGrades = self.ui.lineEdit_31.text()
        SixthPgmStudents = self.ui.lineEdit_32.text()
        SixthPgmPrice = self.ui.lineEdit_33.text()
        SeventhPgmTime = self.ui.comboBox_27.currentText()
        SeventhPgmGrades = self.ui.lineEdit_35.text()
        SeventhPgmStudents = self.ui.lineEdit_36.text()
        SeventhPgmPrice = self.ui.lineEdit_37.text()
        EighthPgmTime = self.ui.comboBox_28.currentText()
        EighthPgmGrades = self.ui.lineEdit_39.text()
        EighthPgmStudents = self.ui.lineEdit_40.text()
        EighthPgmPrice = self.ui.lineEdit_41.text()
        NinthPgmTime = self.ui.comboBox_29.currentText()
        NinthPgmGrades = self.ui.lineEdit_43.text()
        NinthPgmStudents = self.ui.lineEdit_44.text()
        NinthPgmPrice = self.ui.lineEdit_45.text()
        TenthPgmTime = self.ui.comboBox_30.currentText()
        TenthPgmGrades = self.ui.lineEdit_47.text()
        TenthPgmStudents = self.ui.lineEdit_48.text()
        TenthPgmPrice = self.ui.lineEdit_49.text()
        FirstPgmTitle = self.ui.comboBox.currentText()
        SecondPgmTitle = self.ui.comboBox_3.currentText()
        ThirdPgmTitle = self.ui.comboBox_5.currentText()
        FourthPgmTitle = self.ui.comboBox_7.currentText()
        FifthPgmTitle = self.ui.comboBox_9.currentText()
        SixthPgmTitle = self.ui.comboBox_11.currentText()
        SeventhPgmTitle = self.ui.comboBox_13.currentText()
        EighthPgmTitle = self.ui.comboBox_15.currentText()
        NinthPgmTitle = self.ui.comboBox_17.currentText()
        TenthPgmTitle = self.ui.comboBox_19.currentText()
        FirstPgmRoom = str.split(self.ui.comboBox_2.currentText(), ' (')[0]
        SecondPgmRoom = str.split(self.ui.comboBox_4.currentText(), ' (')[0]
        ThirdPgmRoom = str.split(self.ui.comboBox_6.currentText(), ' (')[0]
        FourthPgmRoom = str.split(self.ui.comboBox_8.currentText(), ' (')[0]
        FifthPgmRoom = str.split(self.ui.comboBox_10.currentText(), ' (')[0]
        SixthPgmRoom = str.split(self.ui.comboBox_12.currentText(), ' (')[0]
        SeventhPgmRoom = str.split(self.ui.comboBox_14.currentText(), ' (')[0]
        EighthPgmRoom = str.split(self.ui.comboBox_16.currentText(), ' (')[0]
        NinthPgmRoom = str.split(self.ui.comboBox_18.currentText(), ' (')[0]
        TenthPgmRoom = str.split(self.ui.comboBox_20.currentText(), ' (')[0]
        TotalCost = self.ui.lineEdit_50.text()[1:]
        SpecialConsiderations = self.ui.textEdit.toPlainText()
        PgmTitles = [FirstPgmTitle,SecondPgmTitle,ThirdPgmTitle,FourthPgmTitle,FifthPgmTitle,SixthPgmTitle,
                     SeventhPgmTitle,EighthPgmTitle,NinthPgmTitle,TenthPgmTitle]
        PgmTimes = [FirstPgmTime,SecondPgmTime,ThirdPgmTime,FourthPgmTime,FifthPgmTime,SixthPgmTime,SeventhPgmTime,
                    EighthPgmTime,NinthPgmTime,TenthPgmTime]
        PgmGrades = [FirstPgmGrades,SecondPgmGrades,ThirdPgmGrades,FourthPgmGrades,FifthPgmGrades,SixthPgmGrades,
                     SeventhPgmGrades,EighthPgmGrades,NinthPgmGrades,TenthPgmGrades]
        PgmStudents = [FirstPgmStudents,SecondPgmStudents,ThirdPgmStudents,FourthPgmStudents,FifthPgmStudents,
                       SixthPgmStudents,SeventhPgmStudents,EighthPgmStudents,NinthPgmStudents,TenthPgmStudents]
        PgmRooms = [FirstPgmRoom,SecondPgmRoom,ThirdPgmRoom,FourthPgmRoom,FifthPgmRoom,SixthPgmRoom,SeventhPgmRoom,
                    EighthPgmRoom,NinthPgmRoom,TenthPgmRoom]
        PgmPrices = [FirstPgmPrice,SecondPgmPrice,ThirdPgmPrice,FourthPgmPrice,FifthPgmPrice,SixthPgmPrice,
                     SeventhPgmPrice,EighthPgmPrice,NinthPgmPrice,TenthPgmPrice]
        room_numbers = [number for number in PgmRooms if number != '']
        date_written = datetime.datetime.now().strftime('%m/%d/%Y')

        for index in list(range(len(PgmPrices))):
            if PgmPrices[index] != '':
                PgmPrices[index] = '$'+'{:.2f}'.format(float(PgmPrices[index]))

        if self.contract_field_error_check('new') == 1:
            return

        fileexistsflag = 0
        if os.path.exists(outpath):
            fileexistsflag = 1

        if os.path.exists(path):
            pathflag = 0
        else:
            pathflag = 1

        fieldinfo = {"ContractEMS":ContractEMS,"SchoolName":SchoolName, "SchoolAddress": school_address,
                           "SchoolPhone":SchoolPhone,"SchoolFax":SchoolFax,"ReserverName":ReserverName,
                           "ReserverPhone":ReserverPhone,"ReserverEmail":ReserverEmail,"ContactName":ContactName,
                           "ContactPhone":ContactPhone,"DayDate":PgmDate,"JustDate":JustDate,"JustDay":JustDay,
                           "PayDate":PayDate,"SpecialConsiderations":SpecialConsiderations,"FirstPgmTime":FirstPgmTime,
                           "FirstPgmGrades":FirstPgmGrades,"FirstPgmStudents":FirstPgmStudents,"FirstPgmPrice":PgmPrices[0],
                           "SecondPgmTime":SecondPgmTime,"SecondPgmGrades":SecondPgmGrades,
                           "SecondPgmStudents":SecondPgmStudents,"SecondPgmPrice":PgmPrices[1],"ThirdPgmTime":ThirdPgmTime,
                           "ThirdPgmGrades":ThirdPgmGrades,"ThirdPgmStudents":ThirdPgmStudents,"ThirdPgmPrice":PgmPrices[2],
                           "FourthPgmTime":FourthPgmTime,"FourthPgmGrades":FourthPgmGrades,
                           "FourthPgmStudents":FourthPgmStudents,"FourthPgmPrice":PgmPrices[3],"FifthPgmTime":FifthPgmTime,
                           "FifthPgmGrades":FifthPgmGrades,"FifthPgmStudents":FifthPgmStudents,"FifthPgmPrice":PgmPrices[4],
                           "SixthPgmTime":SixthPgmTime,"SixthPgmGrades":SixthPgmGrades,"SixthPgmStudents":SixthPgmStudents,
                           "SixthPgmPrice":PgmPrices[5],"SeventhPgmTime":SeventhPgmTime,"SeventhPgmGrades":SeventhPgmGrades,
                           "SeventhPgmStudents":SeventhPgmStudents,"SeventhPgmPrice":PgmPrices[6],
                           "EighthPgmTime":EighthPgmTime,"EighthPgmGrades":EighthPgmGrades,
                           "EighthPgmStudents":EighthPgmStudents,"EighthPgmPrice":PgmPrices[7],"NinthPgmTime":NinthPgmTime,
                           "NinthPgmGrades":NinthPgmGrades,"NinthPgmStudents":NinthPgmStudents,"NinthPgmPrice":PgmPrices[8],
                           "TenthPgmTime":TenthPgmTime,"TenthPgmGrades":TenthPgmGrades,"TenthPgmStudents":TenthPgmStudents,
                           "TenthPgmPrice":PgmPrices[9],"FirstPgmTitle":FirstPgmTitle, "SecondPgmTitle":SecondPgmTitle,
                           "ThirdPgmTitle":ThirdPgmTitle, "FourthPgmTitle":FourthPgmTitle, "FifthPgmTitle":FifthPgmTitle,
                           "SixthPgmTitle":SixthPgmTitle, "SeventhPgmTitle":SeventhPgmTitle,
                           "EighthPgmTitle":EighthPgmTitle, "NinthPgmTitle":NinthPgmTitle, "TenthPgmTitle":TenthPgmTitle,
                           "PgmDate": JustDate, "TotalCost": TotalCost, "DateTaken": DateTaken, "BookedBy": BookedBy,
                     "DateWritten": date_written}
        fieldinfo.update({k: None for k in self.contract_info()[5]})  # add payment info keys

        for iteration in range(1):
            root = tkinter.Tk()
            root.withdraw()

            if pathflag == 1:
                tkinter.messagebox.showerror("Error",message="      Unable to access destination file path")
                break

            # Section to write contract spreadsheet
            sheetdate = datetime.datetime.now().strftime("%b%Y")
            currentyear = int(sheetdate[-4:])
            sheetfile = 'inhouse_contract_spreadsheet_'+sheetdate[-4:]+'.xlsx'
            sheetpath = path
            sheettemplate = 'contract_spreadsheet_template.xlsx'

            if not os.path.exists(sheettemplate):
                tkinter.messagebox.showerror("Error!", message="Unable to find spreadsheet template in " +
                                             "specified location")
                break

            list1 = [ContractEMS, SchoolName, school_street, school_city, school_zip,
                     SchoolPhone, ReserverName, ReserverPhone, ContactName, ContactPhone, SchoolFax,
                     SpecialConsiderations, ReserverEmail, JustDate, JustDay, PgmTitles[0], PgmTimes[0],
                     PgmGrades[0], PgmStudents[0], PgmRooms[0], PgmPrices[0], TotalCost, PayDate, BookedBy, '', '']

            # This checks if contract is overwriting a previous one of the same name. If so, the excel record
            # should overwrite the existing excel record instead of appending a new one.
            file_check = [os.path.join(root, name) for root, dirs, files in os.walk(path) for name in files if
                          fieldinfo["ContractEMS"]+'.docx' in name]
            try:
                if not file_check:  # check if file_check is empty, i.e. no matching contract record found
                    if os.path.exists(sheetpath+sheetfile):  # If file for current year exists, open it and get sheets
                        xfile = openpyxl.load_workbook(sheetpath+sheetfile)
                        sheets = xfile.get_sheet_names()
                        lastsheet = sheets[-1]
                        lastsheetyear = int(lastsheet[-4:])

                        if sheetdate in sheets:  # If sheet for current month exists, find bottom and append data after
                                                    # blank row
                            sheet = xfile.get_sheet_by_name(sheetdate)
                            cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q',
                                    'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']

                            checkforblank = 0
                            for i in list(range(1, 1000)):
                                rowdata = []
                                for j in list(range(len(cols))):
                                    rowdata.append(sheet[cols[j]+str(i)].value)
                                if rowdata.count(None) == len(rowdata):
                                    checkforblank += 1
                                    if checkforblank > 1:
                                        row = i
                                        # print('First row for new entry:',row)
                                        break
                                else:
                                    checkforblank = 0

                            i = 0
                            for n in list1:
                                sheet[cols[i]+str(row)] = n
                                i += 1

                            j = row + 1
                            if len(PgmTitles) > 1:
                                for index in list(range(1, len(PgmTitles))):
                                    sheet['P'+str(j)] = PgmTitles[index]
                                    sheet['Q'+str(j)] = PgmTimes[index]
                                    sheet['R'+str(j)] = PgmGrades[index]
                                    sheet['S'+str(j)] = PgmStudents[index]
                                    sheet['T'+str(j)] = PgmRooms[index]
                                    sheet['U'+str(j)] = PgmPrices[index]
                                    j += 1

                            xfile.save(sheetpath+sheetfile)
                        elif lastsheetyear != currentyear:  # Is this never true?!
                            wb = openpyxl.load_workbook(sheettemplate)
                            ws = wb.active
                            ws.title = sheetdate
                            ws.append(list1)
                            wb.save(filename=sheetpath+sheetfile)
                        else:  # If no sheet for current month exists, create it, write title row, and append data
                            ws = xfile.create_sheet(title=sheetdate)
                            firstrow = ['ContractEMS', 'SchoolName', 'SchoolAddress', 'SchoolCity', 'SchoolZIP',
                                        'SchoolPhone', 'ReserverName', 'ReserverPhone', 'ContactName', 'ContactPhone',
                                        'ContactFax', 'SpecialConsiderations', 'ReserverEmail', 'PgmDate', 'PgmDay',
                                        'ProgramTitle', 'PgmTime', 'Grades', 'Students', 'Room', 'CostForPgmHours',
                                        'TOTALCOST', 'DueDate', 'BookedBy', 'Rev or CX', 'Notes']
                            ws.append(firstrow)
                            ws.append(list1)
                            if len(PgmTitles) > 1:
                                i = 3
                                for index in list(range(1, len(PgmTitles))):
                                    ws['P'+str(i)] = PgmTitles[index]
                                    ws['Q'+str(i)] = PgmTimes[index]
                                    ws['R'+str(i)] = PgmGrades[index]
                                    ws['S'+str(i)] = PgmStudents[index]
                                    ws['T'+str(i)] = PgmRooms[index]
                                    ws['U'+str(i)] = PgmPrices[index]
                                    i += 1

                            xfile.save(filename=sheetpath+sheetfile)
                    else:  # If no file for current year exists, create it from template and append data
                        wb = openpyxl.load_workbook(sheettemplate)
                        ws = wb.active
                        ws.title = sheetdate
                        ws.append(list1)
                        if len(PgmTitles) > 1:
                            i = 3
                            for index in list(range(1, len(PgmTitles))):
                                ws['P'+str(i)] = PgmTitles[index]
                                ws['Q'+str(i)] = PgmTimes[index]
                                ws['R'+str(i)] = PgmGrades[index]
                                ws['S'+str(i)] = PgmStudents[index]
                                ws['T'+str(i)] = PgmRooms[index]
                                ws['U'+str(i)] = PgmPrices[index]
                                i += 1

                        wb.save(filename=sheetpath+sheetfile)
                else:  # Case when record number already exists, i.e. must overwrite contract and excel record
                    ok = tkinter.messagebox.askokcancel("Warning", "A contract file with this EMS number "
                                                                   "already exists!\n"
                                                                   "This will overwrite the existing contract."
                                                                   " Continue?")
                    if ok != 1:
                        tkinter.messagebox.showinfo("Aborted", "Contract creation aborted")
                        return
                    contract_kp = self.contract_kp_builder('merge_template.docx', os.path.abspath(file_check[0]))
                    c_date = self.creation_date(os.path.abspath(file_check[0]))  # get written date of existing contract
                    _, found_sheet = self.search_whole_spreadsheet(c_date, contract_kp["ContractEMS"])
                    if found_sheet != c_date:
                        c_date = found_sheet
                    # Check if DateWritten timestamp exists in contract. If so, and if it doesn't point to the correct
                    # spreadsheet page, remove it from the merging dict
                    # FIXME
                    try:
                        if "DateWritten" in contract_kp and contract_kp["DateWritten"]:
                            dw = datetime.datetime.strptime(contract_kp["DateWritten"], '%m/%d/%Y').strftime('%b%Y')
                            if dw == found_sheet:
                                fieldinfo.update({"DateWritten": contract_kp["DateWritten"]})
                    except:
                        pass

                    new_xl_record = self.build_xl_record(fieldinfo, room_numbers, 'not_rev')
                    self.write_xl_record(sheetpath+sheetfile, c_date, fieldinfo["ContractEMS"], new_xl_record)
                    os.remove(os.path.abspath(file_check[0]))
            except PermissionError as err:
                tkinter.messagebox.showerror("Error!", "Permission error. Make sure spreadsheet is not open "
                                                       "anywhere, then try again.\n\n" + str(err))
                return

            # Make sure contract template file exists
            if not os.path.exists(FileName):
                tkinter.messagebox.showerror("Error", message="Unable to find contract template in specified"
                                                              " location.")
                break
            print(date_written)
            # Merge word doc
            try:
                docxmerge(FileName, fieldinfo, outpath)
            except Exception as err:
                tkinter.messagebox.showerror("Error!", str(err))
                return
            tkinter.messagebox.showinfo("Success", message="Contract created and spreadsheet data written " +
                                                           "successfully")


class EntryWidget(tkinter.Tk):
    def __init__(self, entries):
        tkinter.Tk.__init__(self)
        self.title('Email Login')
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry('%dx%d+%d+%d' % (300, 71+17*entries, sw/2-150, sh/2-50))

        l0 = tkinter.Label(self, text="Please enter your email login credentials")
        l1 = tkinter.Label(self, text="Email Address: ")
        l2 = tkinter.Label(self, text="Password: ", height=2)
        l3 = tkinter.Label(self, text="bConnected Key: ", height=2)
        e1 = tkinter.Entry(self, width=30)
        e2 = tkinter.Entry(self, show="\u2022", width=30)
        e3 = tkinter.Entry(self, show="\u2022", width=30)
        b1 = tkinter.Button(self, text="OK", width=20, command=self.ok_button)
        b2 = tkinter.Button(self, text="Cancel", width=20, command=self.cancel_button)

        l0.grid(row=0, columnspan=4)
        if entries == 2:
            l1.grid(row=1, column=0)
            l2.grid(row=2, column=0)
            e1.grid(row=1, column=1, columnspan=3)
            e2.grid(row=2, column=1, columnspan=3)
            b1.grid(row=3, column=0, columnspan=2)
            b2.grid(row=3, column=2, columnspan=2)

        elif entries == 1:
            l3.grid(row=1)
            e3.grid(row=1, column=1, columnspan=3)
            b1.grid(row=2, column=0, columnspan=2)
            b2.grid(row=2, column=2, columnspan=2)

        self.username = e1
        self.password = e2
        self.bc_key = e3
        e1.focus_force()
        self.mainloop()

    def ok_button(self):
        self.withdraw()
        self.quit()
        return self.username.get(), self.password.get(), self.bc_key.get()

    def cancel_button(self):
        self.withdraw()
        self.quit()
        return None, None, None


class CryptoCipher:
    def __init__(self, passphrase):
        passphrase_padded = passphrase + '{' * (32 - len(passphrase))
        passphrase_encoded = b64encode(bytes(passphrase_padded, "utf-8"))
        self.cipher = Fernet(passphrase_encoded)

    def encrypt(self, plain_text):
        return self.cipher.encrypt(bytes(plain_text, "utf-8")).decode()

    def decrypt(self, enc_text):
        return self.cipher.decrypt(bytes(enc_text, "utf-8")).decode()


if __name__ == "__main__":
    app = QtGui.QApplication(sys.argv)
    myapp = MyForm()
    myapp.show()
    sys.exit(app.exec_())
